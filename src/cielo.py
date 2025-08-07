import os
import pandas as pd
import streamlit as st
import logging
from datetime import datetime
from rapidfuzz import fuzz
from openpyxl import load_workbook


# Configura√ß√£o de logging
logging.basicConfig(
    level=logging.DEBUG,  # ou DEBUG para mais detalhes
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("conciliacao.log", encoding="utf-8"),  # grava em arquivo
        logging.StreamHandler()  # mostra no console
    ]
)


# =========================
# Fun√ß√£o de limpeza ERP
# =========================
def limpar_erp(df):
    try:
        with st.spinner("üßπ Limpando dados do ERP..."):
            df["Emiss√£o"] = pd.to_datetime(df["Emiss√£o"], dayfirst=True, errors="coerce")
            parcelas = df["Numero"].str.extract(r"-(\d+)/(\d+)")
            df["Numero da Parcela"] = parcelas[0].astype(float).fillna(1).astype(int)
            df["Total Parcelas"] = parcelas[1].astype(float).fillna(1).astype(int)

            df["Valor"] = (
                df["Valor"].astype(str).str.replace(",", ".", regex=False).astype(float)
            )

    except Exception as e:
        logging.error(f"Erro ao limpar dados ERP: {e}", exc_info=True)
        raise

    return df

# =========================
# Fun√ß√£o de limpeza Cielo
# =========================
def limpar_cielo(df):
    try:
        with st.spinner("üßπ Limpando dados da Cielo..."):
            df = df.iloc[8:].reset_index(drop=True)
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)
            df.columns = df.columns.str.strip().str.lower()

            df = df.rename(columns={
                "valor bruto": "VALOR DA PARCELA",
                "valor l√≠quido": "VALOR L√çQUIDO",
                "n√∫mero da parcela": "PARCELA",
                "quantidade total de parcelas": "TOTAL_PARCELAS",
                "c√≥digo da autoriza√ß√£o": "AUTORIZA√á√ÉO",
                "nsu/doc": "NSU/DOC",
                "data da venda": "DATA DA VENDA",
                "data prevista de pagamento": "DATA DE VENCIMENTO",
                "tipo de lan√ßamento": "TIPO DE LAN√áAMENTO",
            })

            for col in ["VALOR DA PARCELA", "VALOR L√çQUIDO"]:
                df[col] = (
                    df[col].astype(str).str.replace(",", ".", regex=False).astype(float)
                )

            df["PARCELA"] = pd.to_numeric(df["PARCELA"], errors="coerce").fillna(1).astype(int)
            df["TOTAL_PARCELAS"] = pd.to_numeric(df["TOTAL_PARCELAS"], errors="coerce").fillna(1).astype(int)

            for col in ["DATA DA VENDA", "DATA DE VENCIMENTO"]:
                df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")
                
                
            # Mant√©m apenas as colunas mencionadas acima:
            colunas_manter = [
                "VALOR DA PARCELA",
                "VALOR L√çQUIDO",
                "PARCELA",
                "TOTAL_PARCELAS",
                "AUTORIZA√á√ÉO",
                "NSU/DOC",
                "DATA DA VENDA",
                "DATA DE VENCIMENTO",
                "TIPO DE LAN√áAMENTO",
            ]
            df = df[colunas_manter]
    except Exception as e:
        logging.error(f"Erro ao limpar dados Cielo: {e}", exc_info=True)
        raise
    return df





# =========================
# ==Fun√ß√£o de concilia√ß√£o==
# =========================

def conciliar_cielo_erp(df_cielo, df_erp, tolerancia_dias=5, tolerancia_valor=0.20):
    df_cielo = df_cielo.copy()
    df_erp = df_erp.copy()

    # Normalizar chaves
    df_erp["Chave"] = pd.to_numeric(df_erp["Chave"], errors="coerce").astype("Int64")
    df_erp["Usada"] = False

    # Adiciona colunas de resultado na df_cielo
    df_cielo["Autoriza√ß√£o ERP"] = None
    df_cielo["NSU ERP"] = None
    df_cielo["Chave ERP"] = None
    df_cielo["Valor ERP"] = None
    df_cielo["Emiss√£o ERP"] = None
    df_cielo["Parcela ERP"] = None
    df_cielo["Total Parcelas ERP"] = None
    df_cielo["Pessoa do T√≠tulo"] = None
    df_cielo["Pessoa do T√≠tulo"] = None
    df_cielo["Status"] = "N√£o conciliado"
    df_cielo["Pontua√ß√£o"] = 999

    progress_text = st.empty()  # cria um espa√ßo que podemos atualizar
    progress_bar = st.progress(0)
    total = len(df_cielo)

    for i, row in df_cielo.iterrows():
        progresso = (i + 1) / total

        # Atualiza o texto com os registros j√° processados
        progress_text.text(f"üîÑ Conciliando ({i + 1}/{total}) registros...")

        # Atualiza a barra
        progress_bar.progress(progresso)

        # seu processamento da linha aqui

    for i, row in df_cielo.iterrows():
        if pd.isna(row["AUTORIZA√á√ÉO"]) or pd.isna(row["NSU/DOC"]):
            logging.warning(f"‚ö†Ô∏è Linha {i} ignorada por dados ausentes.")
            continue

        logging.debug(f"üîç Linha {i} - Aut: {row['AUTORIZA√á√ÉO']}, NSU: {row['NSU/DOC']}, Parcela: {row['PARCELA']}")

        candidatos = df_erp[
            (~df_erp["Usada"]) &
            (abs((df_erp["Emiss√£o"] - row["DATA DA VENDA"]).dt.days) <= tolerancia_dias) &
            (abs(df_erp["Valor"] - row["VALOR DA PARCELA"]) <= tolerancia_valor) &
            (df_erp["Numero da Parcela"] == row["PARCELA"]) &
            (df_erp["Total Parcelas"] == row["TOTAL_PARCELAS"])
        ]

        logging.debug(f"üîé {len(candidatos)} candidatos encontrados para a linha {i} da Cielo.")

        melhor = None
        menor_pontuacao = float("inf")

        for _, linha in candidatos.iterrows():
            dias_dif = abs((linha["Emiss√£o"] - row["DATA DA VENDA"]).days)
            valor_dif = abs(linha["Valor"] - row["VALOR DA PARCELA"])
            sim_aut = fuzz.ratio(str(linha["Autoriza√ß√£o"]), str(row["AUTORIZA√á√ÉO"]))
            sim_nsu = fuzz.ratio(str(linha["NSU"]), str(row["NSU/DOC"]))

            pontuacao = dias_dif * 10 + valor_dif * 100 + (100 - sim_aut) + (100 - sim_nsu)
            if "Pessoa do T√≠tulo" in linha and linha["Pessoa do T√≠tulo"] != "Cielo":
                    pontuacao += 101

            logging.debug(f"‚û°Ô∏è Testando Chave {linha['Chave']} | Dias: {dias_dif}, Valor: {valor_dif}, Aut: {sim_aut}, NSU: {sim_nsu}, Pontua√ß√£o: {pontuacao:.2f}")

            if pontuacao < menor_pontuacao:
                menor_pontuacao = pontuacao
                melhor = linha

        if melhor is not None:
            idx_erp = df_erp.index[df_erp["Chave"] == melhor["Chave"]].tolist()
            if idx_erp:
                df_erp.at[idx_erp[0], "Usada"] = True

            df_cielo.at[i, "Autoriza√ß√£o ERP"] = melhor["Autoriza√ß√£o"]
            df_cielo.at[i, "NSU ERP"] = melhor["NSU"]
            df_cielo.at[i, "Chave ERP"] = melhor["Chave"]
            df_cielo.at[i, "Valor ERP"] = melhor["Valor"]
            df_cielo.at[i, "Emiss√£o ERP"] = melhor["Emiss√£o"]
            df_cielo.at[i, "Parcela ERP"] = melhor["Numero da Parcela"]
            df_cielo.at[i, "Total Parcelas ERP"] = melhor["Total Parcelas"]
            df_cielo.at[i, "Pessoa do T√≠tulo"] = melhor.get("Pessoa do T√≠tulo", None)
            df_cielo.at[i, "Status"] = "Conciliado"
            df_cielo.at[i, "Pontua√ß√£o"] = round(menor_pontuacao, 0)
            logging.info(f"‚úÖ Linha {i} conciliada com chave {melhor['Chave']} (Pontua√ß√£o: {round(menor_pontuacao, 0)})")
        else:
            logging.info(f"‚ùå Linha {i} n√£o conciliada (sem candidatos adequados)")

    return df_cielo, df_erp 






def main():

    # === BARRA LATERAL ===
    with st.sidebar:
        st.markdown("# App Concilia√ß√£o Banc√°ria")
        st.markdown("### Carregar planilhas")
        caminho_erp = st.file_uploader("ERP (CSV)", type=["csv"], key="erp_uploader")
        caminho_cielo = st.file_uploader("Cielo (XLSX)", type=["xlsx"], key="cielo_uploader")

    # === TELA INICIAL ===
    if caminho_erp is None or caminho_cielo is None:
        st.subheader("Bem-vindo ao Sistema de Concilia√ß√£o")
        st.markdown("""
        <div style='text-align: center; margin-bottom: 20px;'>
            <p>Este sistema realiza a concilia√ß√£o autom√°tica entre:</p>
            <p>‚Ä¢  Cielo</p>
            <p>‚Ä¢ ERP</p>
        </div>
        """, unsafe_allow_html=True)
        st.warning("‚ö†Ô∏è Por favor, fa√ßa upload de ambos os arquivos para iniciar a concilia√ß√£o")
        st.stop()

    def carregar_planilha(caminho):
        if caminho.name.lower().endswith(".csv"):
            return pd.read_csv(caminho, sep=";", encoding="latin1")
        elif caminho.name.lower().endswith(".xlsx") or caminho.name.lower().endswith(".xls"):
            return pd.read_excel(caminho, engine="openpyxl")
        else:
            raise ValueError("‚ùå Formato de arquivo n√£o suportado. S√≥ aceitamos CSV e XLSX.")

    try:
        with st.spinner("üìÇ Carregando planilhas..."):
            df_erp = carregar_planilha(caminho_erp)
            df_cielo = carregar_planilha(caminho_cielo)

        with st.spinner("üîß Iniciando limpeza e concilia√ß√£o dos dados..."):
            df_erp = limpar_erp(df_erp)
            df_cielo = limpar_cielo(df_cielo)
            df_conciliado, df_erp = conciliar_cielo_erp(df_cielo, df_erp)
            df_aba_conciliados = df_conciliado[df_conciliado["Status"] == "Conciliado"].copy()
            df_aba_nao_conciliados = df_conciliado[df_conciliado["Status"] != "Conciliado"].copy()

        totais_conc = {
            "liquido": df_aba_conciliados["VALOR L√çQUIDO"].sum(),
            "parcela": df_aba_conciliados["VALOR DA PARCELA"].sum(),
            "qtd": len(df_aba_conciliados)
        }
        totais_nao = {
            "liquido": df_aba_nao_conciliados["VALOR L√çQUIDO"].sum(),
            "parcela": df_aba_nao_conciliados["VALOR DA PARCELA"].sum(),
            "qtd": len(df_aba_nao_conciliados)
        }

        relatorio_linhas = [
            ["RELAT√ìRIO DE CONCILIA√á√ÉO", "", ""],
            ["CONCILIADO", "", ""],
            ["- Valor L√≠quido Total", "", f"R$ {totais_conc['liquido']:,.2f}"],
            ["- Valor da Parcela Total", "", f"R$ {totais_conc['parcela']:,.2f}"],
            ["- Quantidade de T√≠tulos", "", f"{totais_conc['qtd']}"],
            ["", "", ""],
            ["N√ÉO CONCILIADO", "", ""],
            ["- Valor L√≠quido Total", "", f"R$ {totais_nao['liquido']:,.2f}"],
            ["- Valor da Parcela Total", "", f"R$ {totais_nao['parcela']:,.2f}"],
            ["- Quantidade de T√≠tulos", "", f"{totais_nao['qtd']}"]
        ]
        relatorio_df = pd.DataFrame(relatorio_linhas, columns=["Categoria", "Descri√ß√£o", "Valor"])

        # =====================================================================
        # EXCLUS√ÉO FINAL DAS COLUNAS (AP√ìS TODO O PROCESSAMENTO)
        # =====================================================================
        # Definir colunas a serem exclu√≠das pelos nomes reais
        colunas_para_excluir = [
            "TIPO DE LAN√áAMENTO",   # Coluna I
            "Parcela ERP",          # Coluna O
            "Total Parcelas ERP"    # Coluna P
        ]

        # Aplicar exclus√£o apenas se as colunas existirem
        for col in colunas_para_excluir:
            if col in df_aba_conciliados.columns:
                df_aba_conciliados = df_aba_conciliados.drop(columns=[col])
            if col in df_aba_nao_conciliados.columns:
                df_aba_nao_conciliados = df_aba_nao_conciliados.drop(columns=[col])

        # Agora gerar o Excel com as colunas j√° exclu√≠das


        
        output_path = "Concilia√ß√£o_final.xlsx"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_aba_conciliados.to_excel(writer, sheet_name="Conciliados", index=False)
            df_aba_nao_conciliados.to_excel(writer, sheet_name="N√£o conciliados", index=False)
            relatorio_df.to_excel(writer, sheet_name="Resumo", index=False)

            # Tratar abas especiais (aluguel e estornos) - tamb√©m remover coluna I
            if "TIPO DE LAN√áAMENTO" in df_cielo.columns:
                # Criar c√≥pias para n√£o alterar o original
                df_cielo_sem_coluna = df_cielo.drop(columns=["TIPO DE LAN√áAMENTO"], errors="ignore")
                
                df_aluguel = df_cielo_sem_coluna[df_cielo["TIPO DE LAN√áAMENTO"].str.lower().str.contains("aluguel", na=False)]
                if not df_aluguel.empty:
                    df_aluguel.to_excel(writer, sheet_name="Aluguel de m√°quina", index=False)
                
                df_estornos = df_cielo_sem_coluna[df_cielo["TIPO DE LAN√áAMENTO"].str.lower().str.contains("estorno", na=False)]
                if not df_estornos.empty:
                    df_estornos.to_excel(writer, sheet_name="Estornos", index=False)

        # === INSERIR CHAVES ERP EM BLOCOS NA ABA RESUMO ===
        try:
            wb = load_workbook(output_path)
            ws_conciliados = wb["Conciliados"]
            ws_resumo = wb["Resumo"]

            # Detecta a coluna da Chave ERP
            header = [cell.value for cell in ws_conciliados[1]]
            if "Chave ERP" in header:
                idx_chave = header.index("Chave ERP")
                letra_coluna = chr(65 + idx_chave)

                chaves = [str(cell.value) for cell in ws_conciliados[letra_coluna][1:] if cell.value is not None]

                blocos = [chaves[i:i+2000] for i in range(0, len(chaves), 2000)]
                blocos_concat = [", ".join(bloco) for bloco in blocos]

                start_row = ws_resumo.max_row + 2
                for i, texto in enumerate(blocos_concat, start=1):
                    ws_resumo.cell(row=start_row + i - 1, column=1, value=f"Grupo {i}")
                    ws_resumo.cell(row=start_row + i - 1, column=2, value=texto)

                wb.save(output_path)
            else:
                st.warning("Coluna 'Chave ERP' n√£o encontrada na aba Conciliados")

        except Exception as e:
            st.error(f"‚ùå Erro ao adicionar blocos de Chave ERP: {e}")

        # === INTERFACE FINAL ===
        with st.container():
            st.header("Resultados da Concilia√ß√£o")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("‚úÖ Conciliados", 
                        f"R$ {totais_conc['liquido']:,.2f}", 
                        f"{totais_conc['qtd']} t√≠tulos")
            with col2:
                st.metric("‚ö† N√£o Conciliados", 
                        f"R$ {totais_nao['liquido']:,.2f}", 
                        f"{totais_nao['qtd']} t√≠tulos")

            with st.expander("üìä Ver relat√≥rio completo"):
                st.dataframe(relatorio_df, hide_index=True)

        if os.path.exists(output_path):
            with open(output_path, "rb") as f:
                st.download_button(
                    label="üì• Baixar Planilha de Concilia√ß√£o",
                    data=f,
                    file_name="Concilia√ß√£o_final_cielo.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"‚ùå Erro ao carregar arquivos: {e}")
        st.stop()