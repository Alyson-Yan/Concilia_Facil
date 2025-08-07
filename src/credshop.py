# =========================
#       importações 
# =========================


import io
import os
import logging
import pandas as pd
import streamlit as st
from rapidfuzz import fuzz
from datetime import datetime
from openpyxl import load_workbook
# =========================
# logging de debug
# =========================

logging.basicConfig(
    level=logging.DEBUG,  # ou INFO para menos verbosidade
    format='%(levelname)s:%(message)s'
)



# =========================
# Função de limpeza ERP
# =========================
def limpar_erp(df):
    try:
        with st.spinner("🧹 Limpando dados do ERP..."):
            df["Emissão"] = pd.to_datetime(df["Emissão"], dayfirst=True, errors="coerce")
            parcelas = df["Numero"].str.extract(r"-(\d+)/(\d+)")
            df["Correção"] = pd.to_datetime(df["Correção"], dayfirst=True, errors="coerce")
            df["Numero da Parcela"] = parcelas[0].astype(float).fillna(1).astype(int)
            df["Total Parcelas"] = parcelas[1].astype(float).fillna(1).astype(int)

            df["Valor"] = (
                df["Valor"].astype(str).str.replace(",", ".", regex=False).astype(float)
                )

            # ✅ Tratar a coluna "Taxa": manter somente 2 casas decimais
            if "Taxa" in df.columns:
                df["Taxa"] = df["Taxa"].astype(str).str.replace(",", ".", regex=False)
                df["Taxa"] = df["Taxa"].str.extract(r"(\d+\.\d{1,2})")  # regex para pegar até 2 decimais
                df["Taxa"] = pd.to_numeric(df["Taxa"], errors="coerce")  # converter para float


            # ✅ Excluir colunas indesejadas
            colunas_para_excluir = ["Nome do Cliente", "Tipo", "Carteira", "Caracterização da Venda"]
            df = df.drop(columns=colunas_para_excluir, errors='ignore')

            # ✅ transformar NSU Concentrador em numérico
            df["NSU Concentrador"] = pd.to_numeric(df["NSU Concentrador"], errors="coerce")
            df["NSU"] = pd.to_numeric(df["NSU"], errors="coerce")
        
    except Exception as e:
        logging.error(f"Erro ao limpar dados ERP: {e}", exc_info=True)
        raise

    return df

# ==========================
# função de limpeza CredShop
# ==========================

def limpar_credshop(df):
    try:
            with st.spinner("🧹 Limpando dados da CredShop..."):
                #definindo os cabeçalhos corretos
                CABECALHOS_CREDSHOP = ["Data do Recebimento", "estabelecimento credshop", "pos", "cv", "Tipo de Lançamento", "Data da Venda", "parcela", "Valor Bruto", "Taxa Credshop", "Valor Líquido"
]

            if df.shape[1] == 1: # Verifica se o DataFrame tem apenas uma coluna
                df = df.iloc[:, 0].str.split(",", expand=True) #se tiver apenas uma coluna, divide em várias colunas
                
                df.columns =  CABECALHOS_CREDSHOP  # Aplica os cabeçalhos corretos
                
                df = df.apply(lambda x: x.strip() if isinstance(x, str) else x)  # Remove espaços em branco
                
                # 4. Dividir parcela em duas colunas
                df = df.rename(columns={'parcela': 'parcela_original'})
                df['parcela_original'] = df['parcela_original'].astype(str).str.zfill(4)
                df['parcela'] = df['parcela_original'].str[:2].astype(int)
                df['parcela_total'] = df['parcela_original'].str[2:].astype(int)
                df = df.drop(columns=['parcela_original'])
                
                # Converter colunas de valor para float (substituindo vírgula por ponto)
                colunas_valores = ["Valor Bruto", "Taxa Credshop", "Valor Líquido"]
                df[colunas_valores] = df[colunas_valores].replace(',', '.', regex=True).apply(pd.to_numeric, errors='coerce')
                
                # Converte as datas para datetime (dia primeiro, erros viram NaT)
                df["Data da Venda"] = pd.to_datetime(df["Data da Venda"], dayfirst=True, errors="coerce")
                df["Data do Recebimento"] = pd.to_datetime(df["Data do Recebimento"], dayfirst=True, errors="coerce")

                        # ✅ transformar NSU Concentrador em numérico
                df["cv"] = pd.to_numeric(df["cv"], errors="coerce")
                
                

                
                
    except Exception as e:
        logging.error(f"Erro ao limpar dados CredShop: {e}", exc_info=True)
        raise
    return df



# =====================================
# rename para aplicar conciliador geral
# =====================================
def renomear_colunas_credshop(df_credshop):
    df_credshop.rename(columns={
        "cv": "NSU/DOC",
        "Valor Bruto": "VALOR DA PARCELA",
        "parcela_total": "TOTAL_PARCELAS",
        "parcela": "PARCELA",
    "Data da Venda": "DATA DA VENDA",
    "Valor Líquido": "VALOR LÍQUIDO",
}, inplace=True)
    
    





def conciliar_credshop_erp(df_credshop, df_erp, tolerancia_dias=5, tolerancia_valor=0.20):
    try:
        with st.spinner("🔄 Conciliando CredShop com ERP..."):
            df_credshop = df_credshop.copy()
            df_erp = df_erp.copy()

            # Normalizar chaves
            df_erp["Chave"] = pd.to_numeric(df_erp["Chave"], errors="coerce").astype("Int64")
            df_erp["Usada"] = False

            # Adiciona colunas de resultado na df_credshop
            df_credshop["NSU ERP"] = None
            df_credshop["Chave ERP"] = None
            df_credshop["Valor ERP"] = None
            df_credshop["Emissão ERP"] = None
            df_credshop["Parcela ERP"] = None
            df_credshop["Total Parcelas ERP"] = None
            df_credshop["Pessoa do Título"] = None 
            df_credshop["Status"] = "Não conciliado"
            df_credshop["Pontuação"] = 999


        progress_text = st.empty()  # cria um espaço que podemos atualizar
        progress_bar = st.progress(0)
        total = len(df_credshop)

        for i, row in df_credshop.iterrows():
            progresso = (i + 1) / total

            # Atualiza o texto com os registros já processados
            progress_text.text(f"🔄 Conciliando ({i + 1}/{total}) registros...")

            # Atualiza a barra
            progress_bar.progress(progresso)

            # seu processamento da linha aqui


            for i, row in df_credshop.iterrows():
                if pd.isna(row["NSU/DOC"]):
                    logging.warning(f"⚠️ Linha {i} ignorada por dados ausentes.")
                    continue

                logging.debug(f"🔍 Linha {i} - NSU: {row['NSU/DOC']}, Parcela: {row['PARCELA']}")

                candidatos = df_erp[
                    (~df_erp["Usada"]) &
                    (abs((df_erp["Emissão"] - row["DATA DA VENDA"]).dt.days) <= tolerancia_dias) &
                    (abs(df_erp["Valor"] - row["VALOR DA PARCELA"]) <= tolerancia_valor) &
                    (df_erp["Numero da Parcela"] == row["PARCELA"]) &
                    (df_erp["Total Parcelas"] == row["TOTAL_PARCELAS"])
                ]

                logging.debug(f"🔎 {len(candidatos)} candidatos encontrados para a linha {i} da credshop.")

                melhor = None
                menor_pontuacao = float("inf")

                for _, linha in candidatos.iterrows():
                    dias_dif = abs((linha["Emissão"] - row["DATA DA VENDA"]).days)
                    valor_dif = abs(linha["Valor"] - row["VALOR DA PARCELA"])
                    sim_nsu = fuzz.ratio(str(linha["NSU"]), str(row["NSU/DOC"]))

                    pontuacao = dias_dif * 10 + valor_dif * 100 + (100 - sim_nsu)
                    if "Pessoa do Título" in linha and linha["Pessoa do Título"] != "Credishop":
                        pontuacao += 101

                    logging.debug(f"➡️ Testando Chave {linha['Chave']} | Dias: {dias_dif}, Valor: {valor_dif}, NSU: {sim_nsu}, Pontuação: {pontuacao:.2f}")


                    if pontuacao < menor_pontuacao:
                        menor_pontuacao = pontuacao
                        melhor = linha

                if melhor is not None:
                    idx_erp = df_erp.index[df_erp["Chave"] == melhor["Chave"]].tolist()
                    if idx_erp:
                        df_erp.at[idx_erp[0], "Usada"] = True

                    df_credshop.at[i, "NSU ERP"] = melhor["NSU"]
                    df_credshop.at[i, "Chave ERP"] = melhor["Chave"]
                    df_credshop.at[i, "Valor ERP"] = melhor["Valor"]
                    df_credshop.at[i, "Emissão ERP"] = melhor["Emissão"]
                    df_credshop.at[i, "Parcela ERP"] = melhor["Numero da Parcela"]
                    df_credshop.at[i, "Total Parcelas ERP"] = melhor["Total Parcelas"]
                    df_credshop.at[i, "Pessoa do Título"] = melhor.get("Pessoa do Título", None)
                    df_credshop.at[i, "Status"] = "Conciliado"
                    df_credshop.at[i, "Pontuação"] = round(menor_pontuacao, 0)
                    logging.info(f"✅ Linha {i} conciliada com chave {melhor['Chave']} (Pontuação: {round(menor_pontuacao, 0)})")
            else:
                logging.info(f"❌ Linha {i} não conciliada (sem candidatos adequados)")
    except Exception as e:
        logging.error(f"Erro ao conciliar: {e}", exc_info=True)
        raise
    return df_credshop, df_erp




    # =========================
    #  INTERFACE STREAMLIT
    # =========================
def main():

    #=================
    #==BARRA LATERAL==
    #=================

    with st.sidebar:
        st.markdown("# App Conciliação Bancária")
        st.markdown("### Carregar planilhas")
        caminho_erp = st.file_uploader("ERP (CSV)", type=["csv"], key="erp_uploader")
        caminho_credshop = st.file_uploader("CredShop (CSV)", type=["csv"], key="credshop_uploader")

    #=================
    # AREA PRINCIPAL
    #=================

    if caminho_erp is None or caminho_credshop is None:
        st.subheader("Bem-vindo ao Sistema de Conciliação")
        st.markdown("""
        <div style='text-align: center; margin-bottom: 20px;'>
            <p>Este sistema realiza a conciliação automática entre:</p>
            <p>•  credshop</p>
            <p>• ERP</p>
        </div>
        """, unsafe_allow_html=True)
        st.warning("⚠️ Por favor, faça upload de ambos os arquivos para iniciar a conciliação")
        st.stop()

    def carregar_planilha(caminho, sem_cabecalho=False):
        if caminho.name.lower().endswith(".csv"):
            return pd.read_csv(
                caminho,
                sep=";",
                encoding="latin1",
                header=None if sem_cabecalho else "infer"  # BOOM!
            )
        else:
            raise ValueError("❌ Apenas arquivos CSV são permitidos.")


    try:
        with st.spinner("📂 Carregando planilhas..."):
            df_erp = carregar_planilha(caminho_erp)
            df_credshop = carregar_planilha(caminho_credshop, sem_cabecalho=True)  # força header=None

            with st.spinner("🔧 Iniciando limpeza e conciliação dos dados..."):
                df_erp = limpar_erp(df_erp)
                df_credshop = limpar_credshop(df_credshop)
                renomear_colunas_credshop(df_credshop)
                df_conciliado, df_erp = conciliar_credshop_erp(df_credshop, df_erp)
                df_aba_conciliados = df_conciliado[df_conciliado["Status"] == "Conciliado"].copy()
                df_aba_nao_conciliados = df_conciliado[df_conciliado["Status"] != "Conciliado"].copy()
                # Remover "aluguéis" e "estornos" da aba "Não conciliados"
                if "Tipo de Lançamento" in df_aba_nao_conciliados.columns:
                    tipo_lcto = df_aba_nao_conciliados["Tipo de Lançamento"].str.lower()
                    df_aba_nao_conciliados = df_aba_nao_conciliados[~tipo_lcto.str.contains("aluguel", na=False)]
                    df_aba_nao_conciliados = df_aba_nao_conciliados[~tipo_lcto.str.contains("estorno", na=False)]




        totais_conc = {
            "liquido": df_aba_conciliados["VALOR LÍQUIDO"].sum(),
            "parcela": df_aba_conciliados["VALOR DA PARCELA"].sum(),
            "qtd": len(df_aba_conciliados)
        }
        totais_nao = {
            "liquido": df_aba_nao_conciliados["VALOR LÍQUIDO"].sum(),
            "parcela": df_aba_nao_conciliados["VALOR DA PARCELA"].sum(),
            "qtd": len(df_aba_nao_conciliados)
        }

        relatorio_linhas = [
            ["RELATÓRIO DE CONCILIAÇÃO", "", ""],
            ["CONCILIADO", "", ""],
            ["- Valor Líquido Total", "", f"R$ {totais_conc['liquido']:,.2f}"],
            ["- Valor da Parcela Total", "", f"R$ {totais_conc['parcela']:,.2f}"],
            ["- Quantidade de Títulos", "", f"{totais_conc['qtd']}"],
            ["", "", ""],
            ["NÃO CONCILIADO", "", ""],
            ["- Valor Líquido Total", "", f"R$ {totais_nao['liquido']:,.2f}"],
            ["- Valor da Parcela Total", "", f"R$ {totais_nao['parcela']:,.2f}"],
            ["- Quantidade de Títulos", "", f"{totais_nao['qtd']}"]
        ]
        relatorio_df = pd.DataFrame(relatorio_linhas, columns=["Categoria", "Descrição", "Valor"])

        # =====================================================================
        # EXCLUSÃO FINAL DAS COLUNAS (APÓS TODO O PROCESSAMENTO)
        # =====================================================================
        # Definir colunas a serem excluídas pelos nomes reais
        colunas_para_excluir = [
            "Taxa Credshop",          # Coluna E
            "Total Parcelas ERP",     # Coluna O
            "Parcela ERP",            # Coluna P
            "Emissão ERP",            # Coluna Q
            "Valor ERP"               # Coluna L
        ]
        
        # Aplicar exclusão apenas se as colunas existirem
        for col in colunas_para_excluir:
            if col in df_aba_conciliados.columns:
                df_aba_conciliados = df_aba_conciliados.drop(columns=[col])
            if col in df_aba_nao_conciliados.columns:
                df_aba_nao_conciliados = df_aba_nao_conciliados.drop(columns=[col])
        
        # Agora gerar o Excel com as colunas já excluídas
        output_path = "Conciliação_final.xlsx"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_aba_conciliados.to_excel(writer, sheet_name="Conciliados", index=False)
            df_aba_nao_conciliados.to_excel(writer, sheet_name="Não conciliados", index=False)
            relatorio_df.to_excel(writer, sheet_name="Resumo", index=False)

            if "Tipo de Lançamento" in df_credshop.columns:
                df_credshop["Tipo de Lançamento"] = df_credshop["Tipo de Lançamento"].astype(str)

                df_aluguel = df_credshop[df_credshop["Tipo de Lançamento"].str.lower().str.contains("aluguel", na=False)]
                if not df_aluguel.empty:
                    df_aluguel.to_excel(writer, sheet_name="Aluguel", index=False)

                df_estorno = df_credshop[df_credshop["Tipo de Lançamento"].str.lower().str.contains("estorno", na=False)]
                if not df_estorno.empty:
                    df_estorno.to_excel(writer, sheet_name="Estorno", index=False)

            if "Sheet1" in writer.book.sheetnames:
                writer.book.remove(writer.book["Sheet1"])

        # === INSERE OS BLOCOS DE CHAVE ERP NA ABA RESUMO ===
        try:
            wb = load_workbook(output_path)
            ws_conciliados = wb["Conciliados"]
            ws_resumo = wb["Resumo"]

            # Identifica a coluna "Chave ERP" dinamicamente
            header = [cell.value for cell in ws_conciliados[1]]
            if "Chave ERP" in header:
                idx_chave = header.index("Chave ERP")
                letra_coluna = chr(65 + idx_chave)  # converte índice em letra (A=65)

                # Coleta os valores da coluna usando a letra encontrada
                col_chave = ws_conciliados[letra_coluna]
                chaves = [str(cell.value) for cell in col_chave[1:] if cell.value is not None]
                
                # Divide em blocos de 2000
                blocos = [chaves[i:i+2000] for i in range(0, len(chaves), 2000)]
                blocos_concat = [", ".join(bloco) for bloco in blocos]

                # Adiciona na aba Resumo
                start_row = ws_resumo.max_row + 2
                for i, texto in enumerate(blocos_concat, start=1):
                    ws_resumo.cell(row=start_row + i - 1, column=1, value=f"Grupo {i}")
                    ws_resumo.cell(row=start_row + i - 1, column=2, value=texto)

                wb.save(output_path)
            else:
                st.warning("Coluna 'Chave ERP' não encontrada na aba Conciliados")
        except Exception as e:
            st.error(f"Erro ao adicionar blocos de Chave ERP: {e}")

        # INTERFACE FINAL
        with st.container():
            st.header("Resultados da Conciliação")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("✅ Conciliados", 
                        f"R$ {totais_conc['liquido']:,.2f}", 
                        f"{totais_conc['qtd']} títulos")
            with col2:
                st.metric("⚠ Não Conciliados", 
                        f"R$ {totais_nao['liquido']:,.2f}", 
                        f"{totais_nao['qtd']} títulos")

            with st.expander("📊 Ver relatório completo"):
                st.dataframe(relatorio_df, hide_index=True)

        if os.path.exists(output_path):
            with open(output_path, "rb") as f:
                st.download_button(
                    label="📥 Baixar Planilha de Conciliação",
                    data=f,
                    file_name="Conciliação_final_credshop.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"❌ Erro ao carregar arquivos: {e}")
        st.stop()

if __name__ == "__main__":
    main()