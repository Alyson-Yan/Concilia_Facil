#CIELO FUNCIONANDO - Ultima atualização: 16/04/2026
#____________________________________________________________________________________________________________________________________________________________________________________________________________________________
import os
import pandas as pd
import streamlit as st
import logging
from datetime import datetime
from rapidfuzz import fuzz
from openpyxl import load_workbook


# Configuração de logging
logging.basicConfig(
    level=logging.DEBUG,  # ou DEBUG para mais detalhes
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("conciliacao.log", encoding="utf-8"),  # grava em arquivo
        logging.StreamHandler()  # mostra no console
    ]
)


def limpar_erp(df):
    try:
        with st.spinner("🧹 Limpando dados do ERP..."):

            # =========================
            # 1. DERIVAÇÕES (ANTES DO RENAME)
            # =========================
            parcelas = df["Numero"].str.extract(r"-(\d+)/(\d+)")
            df["Numero da Parcela"] = parcelas[0].astype(float).fillna(1).astype(int)
            df["Total Parcelas"] = parcelas[1].astype(float).fillna(1).astype(int)

            # =========================
            # 2. RENOMEAÇÃO (PADRÃO FINAL)
            # =========================
            df = df.rename(columns={
                "1o. Agrupamento": "AGRUPAMENTO ERP",
                "Ch Criação": "CHAVE DE CRIAÇÃO ERP",
                "Chave": "CHAVE ERP",
                "Nome do Cliente": "NOME DO CLIENTE ERP",
                "Tipo": "TIPO DE LANÇAMENTO ERP",
                "Carteira": "CARTEIRA ERP",
                "Numero": "NUMERO ERP",
                "Caracterização da Venda": "CARACTERIZAÇÃO DA VENDA ERP",
                "NSU": "NSU ERP",
                "NSU Concentrador": "NSU CONCENTRADOR ERP",
                "Autorização": "AUTORIZAÇÃO ERP",
                "Emissão": "EMISSÃO ERP",
                "Correção": "CORREÇÃO ERP",
                "Valor": "VALOR BRUTO ERP",
                "Vr Corrigido": "VALOR LIQUIDO ERP",
                "Taxa": "TAXA ERP",
            })

            # =========================
            # 3. CONVERSÕES FINAIS
            # =========================
            df["EMISSÃO ERP"] = pd.to_datetime(df["EMISSÃO ERP"], dayfirst=True, errors="coerce")

            df["VALOR BRUTO ERP"] = (
                df["VALOR BRUTO ERP"].astype(str).str.replace(",", ".", regex=False).astype(float)
            )

    except Exception as e:
        logging.error(f"Erro ao limpar dados ERP: {e}", exc_info=True)
        raise

    return df

# =========================
# Função de limpeza Cielo
# =========================
def limpar_cielo(df):
    try:
        with st.spinner("🧹 Limpando dados da Cielo..."):
            df = df.iloc[8:].reset_index(drop=True)
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)
            df.columns = df.columns.str.strip().str.lower()

            df = df.rename(columns={
                "valor bruto":"VALOR BRUTO CIELO",
                "valor líquido":"VALOR LÍQUIDO CIELO",
                "data da venda":"DATA DA VENDA CIELO",
                "data prevista de pagamento":"DATA DE VENCIMENTO CIELO",
                "código da autorização":"AUTORIZAÇÃO CIELO",
                "nsu/doc":"NSU/DOC CIELO",
                "número da parcela":"NUMERO DA PARCELA CIELO",
                "tipo de lançamento": "TIPO DE LANÇAMENTO CIELO",
                "quantidade total de parcelas":"TOTAL PARCELAS CIELO",
            })

            for col in ["VALOR BRUTO CIELO", "VALOR LÍQUIDO CIELO"]:
                df[col] = (
                    df[col].astype(str).str.replace(",", ".", regex=False).astype(float)
                )

            df["NUMERO DA PARCELA CIELO"] = pd.to_numeric(df["NUMERO DA PARCELA CIELO"], errors="coerce").fillna(1).astype(int)
            df["TOTAL PARCELAS CIELO"] = pd.to_numeric(df["TOTAL PARCELAS CIELO"], errors="coerce").fillna(1).astype(int)

            for col in ["DATA DA VENDA CIELO", "DATA DE VENCIMENTO CIELO"]:
                df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")
                
                            # Mantém apenas as colunas mencionadas acima:
            colunas_manter = [
                "VALOR BRUTO CIELO",
                "VALOR LÍQUIDO CIELO",
                "DATA DA VENDA CIELO",
                "DATA DE VENCIMENTO CIELO",
                "AUTORIZAÇÃO CIELO",
                "NSU/DOC CIELO",
                "NUMERO DA PARCELA CIELO",
                "TIPO DE LANÇAMENTO CIELO",
                "TOTAL PARCELAS CIELO",
                ]
            df = df[colunas_manter]
    except Exception as e:
        logging.error(f"Erro ao limpar dados Cielo: {e}", exc_info=True)
        raise
    return df





# =========================
# ==Função de conciliação==
# =========================

def conciliar_cielo_erp(df_cielo, df_erp, tolerancia_dias=5, tolerancia_valor=0.20):
    df_cielo = df_cielo.copy()
    df_erp = df_erp.copy()

    # Normalização de chave
    df_erp["CHAVE ERP"] = pd.to_numeric(df_erp["CHAVE ERP"], errors="coerce").astype("string")
    df_erp["Usada"] = False

    # Colunas de saída
    cols_saida = [
        "AUTORIZAÇÃO ERP",
        "NSU ERP",
        "CHAVE ERP",
        "VALOR BRUTO ERP",
        "VALOR LIQUIDO ERP",
        "EMISSÃO ERP",
        "Parcela ERP",
        "Total Parcelas ERP",
    ]
    for c in cols_saida:
        df_cielo[c] = None

    df_cielo["Status"] = "Não conciliado"
    df_cielo["Pontuação"] = 999

    progress_text = st.empty()
    progress_bar = st.progress(0)
    total = len(df_cielo)

    for i, row in df_cielo.iterrows():
        progress_text.text(f"🔄 Conciliando ({i + 1}/{total}) registros...")
        progress_bar.progress((i + 1) / total)

        if pd.isna(row["AUTORIZAÇÃO CIELO"]) or pd.isna(row["NSU/DOC CIELO"]):
            df_cielo.at[i, "Status"] = "Não conciliado"
            continue

        candidatos = df_erp[
            (~df_erp["Usada"]) &
            (abs((df_erp["EMISSÃO ERP"] - row["DATA DA VENDA CIELO"]).dt.days) <= tolerancia_dias) &
            (abs(df_erp["VALOR BRUTO ERP"] - row["VALOR BRUTO CIELO"]) <= tolerancia_valor) &
            (df_erp["Numero da Parcela"] == row["NUMERO DA PARCELA CIELO"]) &
            (df_erp["Total Parcelas"] == row["TOTAL PARCELAS CIELO"])
        ]

        melhor = None
        menor_pontuacao = float("inf")

        for _, linha in candidatos.iterrows():
            dias_dif = abs((linha["EMISSÃO ERP"] - row["DATA DA VENDA CIELO"]).days)
            valor_dif = abs(linha["VALOR BRUTO ERP"] - row["VALOR BRUTO CIELO"])

            sim_aut = fuzz.ratio(str(linha["AUTORIZAÇÃO ERP"]), str(row["AUTORIZAÇÃO CIELO"]))
            sim_nsu = fuzz.ratio(str(linha["NSU ERP"]), str(row["NSU/DOC CIELO"]))

            pontuacao = dias_dif * 10 + valor_dif * 100 + (100 - sim_aut) + (100 - sim_nsu)

            if pontuacao < menor_pontuacao:
                menor_pontuacao = pontuacao
                melhor = linha

        if melhor is None:
            df_cielo.at[i, "Status"] = "Não conciliado"
            continue

        # Marca uso ERP
        idx = df_erp.index[df_erp["CHAVE ERP"] == melhor["CHAVE ERP"]]
        if len(idx) > 0:
            df_erp.at[idx[0], "Usada"] = True

        # Divergências
        dias_dif = abs((melhor["EMISSÃO ERP"] - row["DATA DA VENDA CIELO"]).days)
        valor_dif = abs(melhor["VALOR BRUTO ERP"] - row["VALOR BRUTO CIELO"])

        # 🔥 DETECÇÃO ESTRUTURAL (CRÍTICO)
        nsu_dif = str(melhor["NSU ERP"]).lstrip("0") != str(row["NSU/DOC CIELO"]).lstrip("0")
        aut_dif = str(melhor["AUTORIZAÇÃO ERP"]).strip() != str(row["AUTORIZAÇÃO CIELO"]).strip()

        divergencias = []

        if valor_dif > 0:
            divergencias.append("Valor divergente")

        if dias_dif > 0:
            divergencias.append("Data divergente")

        if nsu_dif:
            divergencias.append("NSU divergente")

        if aut_dif:
            divergencias.append("Autorização divergente")

        # Status final
        if len(divergencias) == 0:
            status = "Conciliado"
        else:
            status = "Conciliado com: " + " | ".join(divergencias)

        # preenchimento
        df_cielo.at[i, "AUTORIZAÇÃO ERP"] = melhor["AUTORIZAÇÃO ERP"]
        df_cielo.at[i, "NSU ERP"] = melhor["NSU ERP"]
        df_cielo.at[i, "CHAVE ERP"] = melhor["CHAVE ERP"]
        df_cielo.at[i, "VALOR BRUTO ERP"] = melhor["VALOR BRUTO ERP"]
        df_cielo.at[i, "VALOR LIQUIDO ERP"] = melhor.get("VALOR LIQUIDO ERP", None)
        df_cielo.at[i, "EMISSÃO ERP"] = melhor["EMISSÃO ERP"]
        df_cielo.at[i, "Status"] = status
        df_cielo.at[i, "Pontuação"] = round(menor_pontuacao, 2)

    return df_cielo, df_erp






def main():

    # === BARRA LATERAL ===
    with st.sidebar:
        st.markdown("# App Conciliação Bancária")
        st.markdown("### Carregar planilhas")
        caminho_erp = st.file_uploader("ERP (CSV)", type=["csv"], key="erp_uploader")
        caminho_cielo = st.file_uploader("Cielo (XLSX)", type=["xlsx"], key="cielo_uploader")

    # === TELA INICIAL ===
    if caminho_erp is None or caminho_cielo is None:
        st.subheader("Bem-vindo ao Sistema de Conciliação")
        st.markdown("""
        <div style='text-align: center; margin-bottom: 20px;'>
            <p>Este sistema realiza a conciliação automática entre:</p>
            <p>•  Cielo</p>
            <p>• ERP</p>
        </div>
        """, unsafe_allow_html=True)
        st.warning("⚠️ Por favor, faça upload de ambos os arquivos para iniciar a conciliação")
        st.stop()

    def carregar_planilha(caminho):
        if caminho.name.lower().endswith(".csv"):
            return pd.read_csv(caminho, sep=";", encoding="latin1")
        elif caminho.name.lower().endswith(".xlsx") or caminho.name.lower().endswith(".xls"):
            return pd.read_excel(caminho, engine="openpyxl")
        else:
            raise ValueError("❌ Formato de arquivo não suportado. Só aceitamos CSV e XLSX.")

    try:
        with st.spinner("📂 Carregando planilhas..."):
            df_erp = carregar_planilha(caminho_erp)
            df_cielo = carregar_planilha(caminho_cielo)

        with st.spinner("🔧 Iniciando limpeza e conciliação dos dados..."):
            df_erp = limpar_erp(df_erp)
            df_cielo = limpar_cielo(df_cielo)
            df_conciliado, df_erp = conciliar_cielo_erp(df_cielo, df_erp)
            df_aba_conciliados = df_conciliado[
            df_conciliado["Status"].str.startswith("Conciliado", na=False)].copy()
            df_aba_conciliados = df_conciliado[
        df_conciliado["Status"].str.startswith("Conciliado", na=False)].copy()
        df_aba_nao_conciliados = df_conciliado[~df_conciliado["Status"].str.startswith("Conciliado", na=False)].copy()

        totais_conc = {
            "liquido": df_aba_conciliados["VALOR LÍQUIDO CIELO"].sum(),
            "parcela": df_aba_conciliados["VALOR BRUTO CIELO"].sum(),
            "qtd": len(df_aba_conciliados)
        }
        totais_nao = {
            "liquido": df_aba_nao_conciliados["VALOR LÍQUIDO CIELO"].sum(),
            "parcela": df_aba_nao_conciliados["VALOR BRUTO CIELO"].sum(),
            "qtd": len(df_aba_nao_conciliados)
        }

        relatorio_linhas = [
            ["RELATÓRIO DE CONCILIAÇÃO", "", ""],
            ["CONCILIADO", "", ""],
            ["- Valor Líquido Total", "", f"R$ {totais_conc['liquido']:,.2f}"],
            ["- Valor da Parcela Total", "", f"R$ {totais_conc['parcela']:,.2f}"],
            ["- Quantidade de Títulos", "", f"{totais_conc['qtd']}"],
            ["", "", ""],
            ["NÃO CONCILIADO", "", ""],
            ["- Valor Líquido Total", "", f"R$ {totais_nao['liquido']:,.2f}"],
            ["- Valor da Parcela Total", "", f"R$ {totais_nao['parcela']:,.2f}"],
            ["- Quantidade de Títulos", "", f"{totais_nao['qtd']}"]
        ]
        relatorio_df = pd.DataFrame(relatorio_linhas, columns=["Categoria", "Descrição", "VALOR BRUTO ERP"])

        # =====================================================================
        # EXCLUSÃO FINAL DAS COLUNAS (APÓS TODO O PROCESSAMENTO)
        # =====================================================================
        # Definir colunas a serem excluídas pelos nomes reais
        colunas_para_excluir = [
            "TIPO DE LANÇAMENTO CIELO",   # Coluna I
            "Parcela ERP",          # Coluna O
            "Total Parcelas ERP"    # Coluna P
        ]

        # Aplicar exclusão apenas se as colunas existirem
        for col in colunas_para_excluir:
            if col in df_aba_conciliados.columns:
                df_aba_conciliados = df_aba_conciliados.drop(columns=[col])
            if col in df_aba_nao_conciliados.columns:
                df_aba_nao_conciliados = df_aba_nao_conciliados.drop(columns=[col])

        # Agora gerar o Excel com as colunas já excluídas


        
        output_path = "Conciliação_final.xlsx"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_aba_conciliados.to_excel(writer, sheet_name="Conciliados", index=False)
            df_aba_nao_conciliados.to_excel(writer, sheet_name="Não conciliados", index=False)
            relatorio_df.to_excel(writer, sheet_name="Resumo", index=False)

            # Tratar abas especiais (aluguel e estornos) - também remover coluna I
            if "TIPO DE LANÇAMENTO CIELO" in df_cielo.columns:
                # Criar cópias para não alterar o original
                df_cielo_sem_coluna = df_cielo.drop(columns=["TIPO DE LANÇAMENTO CIELO"], errors="ignore")
                
                df_aluguel = df_cielo_sem_coluna[df_cielo["TIPO DE LANÇAMENTO CIELO"].str.lower().str.contains("aluguel", na=False)]
                if not df_aluguel.empty:
                    df_aluguel.to_excel(writer, sheet_name="Aluguel de máquina", index=False)
                
                df_estornos = df_cielo_sem_coluna[df_cielo["TIPO DE LANÇAMENTO CIELO"].str.lower().str.contains("estorno", na=False)]
                if not df_estornos.empty:
                    df_estornos.to_excel(writer, sheet_name="Estornos", index=False)

        # === INSERIR CHAVES ERP EM BLOCOS NA ABA RESUMO ===
        try:
            wb = load_workbook(output_path)
            ws_conciliados = wb["Conciliados"]
            ws_resumo = wb["Resumo"]

            # Detecta a coluna da Chave ERP
            header = [cell.value for cell in ws_conciliados[1]]
            if "CHAVE ERP" in header:
                idx_chave = header.index("CHAVE ERP")
                letra_coluna = chr(65 + idx_chave)

                chaves = [str(cell.value) for cell in ws_conciliados[letra_coluna][1:] if cell.value is not None]

                blocos = [chaves[i:i+2000] for i in range(0, len(chaves), 2000)]
                blocos_concat = [", ".join(bloco) for bloco in blocos]

                start_row = ws_resumo.max_row + 2
                for i, texto in enumerate(blocos_concat, start=1):
                    ws_resumo.cell(row=start_row + i - 1, column=1, value=f"Grupo {i}")
                    ws_resumo.cell(row=start_row + i - 1, column=2, value=texto)

                wb.save(output_path)
            else:
                st.warning("Coluna 'Chave ERP' não encontrada na aba Conciliados")

        except Exception as e:
            st.error(f"❌ Erro ao adicionar blocos de Chave ERP: {e}")

        # === INTERFACE FINAL ===
        with st.container():
            st.header("Resultados da Conciliação")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("✅ Conciliados", 
                        f"R$ {totais_conc['liquido']:,.2f}", 
                        f"{totais_conc['qtd']} títulos")
            with col2:
                st.metric("⚠ Não Conciliados", 
                        f"R$ {totais_nao['liquido']:,.2f}", 
                        f"{totais_nao['qtd']} títulos")

            with st.expander("📊 Ver relatório completo"):
                st.dataframe(relatorio_df, hide_index=True)

        if os.path.exists(output_path):
            with open(output_path, "rb") as f:
                st.download_button(
                    label="📥 Baixar Planilha Final",
                    data=f,
                    file_name="Conciliação_final_cielo.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"❌ Erro ao carregar arquivos: {e}")
        st.stop()