# =========================
#       importações 
# =========================


import io
import os
import logging
import pandas as pd
import streamlit as st
from rapidfuzz import fuzz
from datetime import datetime
from openpyxl import load_workbook
# =========================
# logging de debug
# =========================

logging.basicConfig(
    level=logging.DEBUG,  # ou INFO para menos verbosidade
    format='%(levelname)s:%(message)s'
)



def limpar_erp(df):
    try:
        with st.spinner("🧹 Limpando dados do ERP..."):

            # =========================
            # 1. DERIVAÇÕES (ANTES DO RENAME)
            # =========================
            parcelas = df["Numero"].str.extract(r"-(\d+)/(\d+)")
            df["Numero da Parcela"] = parcelas[0].astype(float).fillna(1).astype(int)
            df["Total Parcelas"] = parcelas[1].astype(float).fillna(1).astype(int)

            # =========================
            # 2. RENOMEAÇÃO (PADRÃO FINAL)
            # =========================
            df = df.rename(columns={
                "1o. Agrupamento": "AGRUPAMENTO ERP",
                "Ch Criação": "CHAVE DE CRIAÇÃO ERP",
                "Chave": "CHAVE ERP",
                "Nome do Cliente": "NOME DO CLIENTE ERP",
                "Tipo": "TIPO DE LANÇAMENTO ERP",
                "Carteira": "CARTEIRA ERP",
                "Numero": "NUMERO ERP",
                "Caracterização da Venda": "CARACTERIZAÇÃO DA VENDA ERP",
                "NSU": "NSU ERP",
                "NSU Concentrador": "NSU CONCENTRADOR ERP",
                "Autorização": "AUTORIZAÇÃO ERP",
                "Emissão": "EMISSÃO ERP",
                "Correção": "CORREÇÃO ERP",
                "Valor": "VALOR BRUTO ERP",
                "Vr Corrigido": "VALOR LIQUIDO ERP",
                "Taxa": "TAXA ERP",
            })

            # =========================
            # 3. CONVERSÕES
            # =========================
            df["EMISSÃO ERP"] = pd.to_datetime(df["EMISSÃO ERP"], dayfirst=True, errors="coerce")

            df["VALOR BRUTO ERP"] = (
                df["VALOR BRUTO ERP"]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .astype(float)
            )

            # =========================
            # 4. LIMPEZA DO NSU
            # =========================
            if "NSU ERP" in df.columns:
                df["NSU ERP"] = (
                    df["NSU ERP"]
                    .astype(str)
                    .str.strip()              # remove espaços
                    .str.lstrip("0")         # remove zeros à esquerda
                    .replace("", pd.NA)      # evita vazio
                )

    except Exception as e:
        logging.error(f"Erro ao limpar dados ERP: {e}", exc_info=True)
        raise

    return df

# ==========================
# função de limpeza CredShop
# ==========================

def limpar_credshop(df):
    try:
            with st.spinner("🧹 Limpando dados da CredShop..."):
                #definindo os cabeçalhos corretos
                CABECALHOS_CREDSHOP = ["DATA DO RECEBIMENTO CREDISHOP","ESTABELECIMENTO CREDISHOP", "POS CREDISHOP", "NSU/DOC CREDISHOP", "TIPO DE LANÇAMENTO CREDISHOP", "DATA DE VENDA CREDISHOP", "parcela", "VALOR BRUTO CREDISHOP", "TAXA CREDISHOP", "VALOR LIQUIDO CREDISHOP"
]

            if df.shape[1] == 1: # Verifica se o DataFrame tem apenas uma coluna
                df = df.iloc[:, 0].str.split(",", expand=True) #se tiver apenas uma coluna, divide em várias colunas
                
                df.columns =  CABECALHOS_CREDSHOP  # Aplica os cabeçalhos corretos
                
                df = df.apply(lambda x: x.strip() if isinstance(x, str) else x)  # Remove espaços em branco
                
                # 4. Dividir parcela em duas colunas
                df = df.rename(columns={'parcela': 'parcela_original'})
                df['parcela_original'] = df['parcela_original'].astype(str).str.zfill(4)
                df['parcela'] = df['parcela_original'].str[:2].astype(int)
                df['parcela_total'] = df['parcela_original'].str[2:].astype(int)
                df = df.drop(columns=['parcela_original'])
                
                # Converter colunas de valor para float (substituindo vírgula por ponto)
                colunas_valores = ["VALOR BRUTO CREDISHOP", "TAXA CREDISHOP", "VALOR LIQUIDO CREDISHOP"]
                df[colunas_valores] = df[colunas_valores].replace(',', '.', regex=True).apply(pd.to_numeric, errors='coerce')
                
                # Converte as datas para datetime (dia primeiro, erros viram NaT)
                df["DATA DE VENDA CREDISHOP"] = pd.to_datetime(df["DATA DE VENDA CREDISHOP"], dayfirst=True, errors="coerce")
                df["DATA DO RECEBIMENTO CREDISHOP"] = pd.to_datetime(df["DATA DO RECEBIMENTO CREDISHOP"], dayfirst=True, errors="coerce")

                        # ✅ transformar NSU Concentrador em numérico
                df["NSU/DOC CREDISHOP"] = pd.to_numeric(df["NSU/DOC CREDISHOP"], errors="coerce")
                

                
    except Exception as e:
        logging.error(f"Erro ao limpar dados CredShop: {e}", exc_info=True)
        raise
    return df



# =====================================
# rename para aplicar conciliador geral
# =====================================
def renomear_colunas_credshop(df_credshop):
    df_credshop.rename(columns={
        "parcela_total": "TOTAL PARCELA CREDISHOP",
        "parcela": "PARCELA ATUAL CREDISHOP",
}, inplace=True)
    
    





def conciliar_credshop_erp(df_credshop, df_erp, tolerancia_dias=5, tolerancia_valor=0.20):
    try:
        with st.spinner("🔄 Conciliando CredShop com ERP..."):
            df_credshop = df_credshop.copy()
            df_erp = df_erp.copy()

            # Normalizar chave
            df_erp["CHAVE ERP"] = pd.to_numeric(df_erp["CHAVE ERP"], errors="coerce").astype("string")
            df_erp["Usada"] = False

            # Colunas resultado
            df_credshop["NSU ERP"] = None
            df_credshop["CHAVE ERP"] = None
            df_credshop["VALOR BRUTO ERP"] = None
            df_credshop["EMISSÃO ERP"] = None
            df_credshop["NUMERO DA PARCELA ERP"] = None
            df_credshop["TOTAL PARCELAS ERP"] = None
            df_credshop["Pessoa do Título"] = None

            df_credshop["DIF_DIAS"] = None
            df_credshop["DIF_VALOR"] = None
            df_credshop["Status"] = "Não conciliado"
            df_credshop["Pontuação"] = 999

        progress_text = st.empty()
        progress_bar = st.progress(0)
        total = len(df_credshop)

        for i, row in df_credshop.iterrows():
            progresso = (i + 1) / total
            progress_text.text(f"🔄 Conciliando ({i + 1}/{total}) registros...")
            progress_bar.progress(progresso)

            if pd.isna(row["NSU/DOC CREDISHOP"]):
                continue

            candidatos = df_erp[
                (~df_erp["Usada"]) &
                (abs((df_erp["EMISSÃO ERP"] - row["DATA DE VENDA CREDISHOP"]).dt.days) <= tolerancia_dias) &
                (abs(df_erp["VALOR BRUTO ERP"] - row["VALOR BRUTO CREDISHOP"]) <= tolerancia_valor) &
                (df_erp["Numero da Parcela"] == row["PARCELA ATUAL CREDISHOP"]) &
                (df_erp["Total Parcelas"] == row["TOTAL PARCELA CREDISHOP"])
            ]

            melhor = None
            menor_pontuacao = float("inf")
            melhor_status = "Não conciliado"
            melhor_dif_dias = None
            melhor_dif_valor = None

            for _, linha in candidatos.iterrows():
                dias_dif = abs((linha["EMISSÃO ERP"] - row["DATA DE VENDA CREDISHOP"]).days)
                valor_dif = abs(linha["VALOR BRUTO ERP"] - row["VALOR BRUTO CREDISHOP"])
                sim_nsu = fuzz.ratio(str(linha["NSU ERP"]), str(row["NSU/DOC CREDISHOP"]))

                # SCORE BASE
                dias_dif = min(dias_dif, 30) if pd.notna(dias_dif) else 30
                valor_dif = min(valor_dif, 1000) if pd.notna(valor_dif) else 1000
                sim_nsu = sim_nsu if pd.notna(sim_nsu) else 0

                # Score controlado (igual filosofia Santander)
                pontuacao = (
                    dias_dif * 10 +
                    valor_dif * 100 +
                    (100 - sim_nsu)
                )

                # HARD LIMIT (anti-bomba nuclear)
                pontuacao = min(pontuacao, 999)

                if "Pessoa do Título" in linha and linha["Pessoa do Título"] != "Credishop":
                    pontuacao += 101

                # =========================
                # STATUS DETALHADO (PADRÃO SANTANDER)
                # =========================
                status_lista = ["Conciliado"]

                # 🔥 QUALQUER DIFERENÇA DE VALOR
                if valor_dif != 0:
                    status_lista.append("Divergência de Valor")

                # 🔥 QUALQUER DIFERENÇA DE DATA
                if dias_dif != 0:
                    status_lista.append("Divergência de Data")

                # 🔥 PARCELA
                if row["PARCELA ATUAL CREDISHOP"] != linha["Numero da Parcela"]:
                    status_lista.append("Divergência de Parcela")

                # 🔥 TOTAL PARCELAS
                if row["TOTAL PARCELA CREDISHOP"] != linha["Total Parcelas"]:
                    status_lista.append("Divergência de Total de Parcelas")

                # 🔥 NSU (igual lógica Santander)
                nsu_cred = str(row["NSU/DOC CREDISHOP"]).strip()
                nsu_erp = str(linha["NSU ERP"]).strip()

                if nsu_cred != nsu_erp:
                    status_lista.append("Divergência de NSU")

                status_final = " e ".join(status_lista) if len(status_lista) > 1 else "Conciliado"

                # Escolhe melhor
                if pontuacao < menor_pontuacao:
                    menor_pontuacao = pontuacao
                    melhor = linha
                    melhor_status = status_final
                    melhor_dif_dias = dias_dif
                    melhor_dif_valor = valor_dif

            # =========================
            # APLICA RESULTADO
            # =========================
            if melhor is not None:
                idx_erp = df_erp.index[df_erp["CHAVE ERP"] == melhor["CHAVE ERP"]].tolist()
                if idx_erp:
                    df_erp.at[idx_erp[0], "Usada"] = True

                df_credshop.at[i, "NSU ERP"] = melhor["NSU ERP"]
                df_credshop.at[i, "CHAVE ERP"] = melhor["CHAVE ERP"]
                df_credshop.at[i, "VALOR BRUTO ERP"] = melhor["VALOR BRUTO ERP"]
                df_credshop.at[i, "EMISSÃO ERP"] = melhor["EMISSÃO ERP"]
                df_credshop.at[i, "NUMERO DA PARCELA ERP"] = melhor["Numero da Parcela"]
                df_credshop.at[i, "TOTAL PARCELAS ERP"] = melhor["Total Parcelas"]
                df_credshop.at[i, "Pessoa do Título"] = melhor.get("Pessoa do Título", None)

                df_credshop.at[i, "Status"] = melhor_status
                df_credshop.at[i, "Pontuação"] = round(menor_pontuacao, 0)
                df_credshop.at[i, "DIF_DIAS"] = melhor_dif_dias
                df_credshop.at[i, "DIF_VALOR"] = melhor_dif_valor

            else:
                df_credshop.at[i, "Status"] = "Não conciliado"
                df_credshop.at[i, "Pontuação"] = 999

    except Exception as e:
        logging.error(f"Erro ao conciliar: {e}", exc_info=True)
        raise

    return df_credshop, df_erp


    # =========================
    #  INTERFACE STREAMLIT
    # =========================
def main():

    #=================
    #==BARRA LATERAL==
    #=================

    with st.sidebar:
        st.markdown("# App Conciliação Bancária")
        st.markdown("### Carregar planilhas")
        caminho_erp = st.file_uploader("ERP (CSV)", type=["csv"], key="erp_uploader")
        caminho_credshop = st.file_uploader("CredShop (CSV)", type=["csv"], key="credshop_uploader")

    #=================
    # AREA PRINCIPAL
    #=================

    if caminho_erp is None or caminho_credshop is None:
        st.subheader("Bem-vindo ao Sistema de Conciliação")
        st.markdown("""
        <div style='text-align: center; margin-bottom: 20px;'>
            <p>Este sistema realiza a conciliação automática entre:</p>
            <p>•  credshop</p>
            <p>• ERP</p>
        </div>
        """, unsafe_allow_html=True)
        st.warning("⚠️ Por favor, faça upload de ambos os arquivos para iniciar a conciliação")
        st.stop()

    def carregar_planilha(caminho, sem_cabecalho=False):
        if caminho.name.lower().endswith(".csv"):
            return pd.read_csv(
                caminho,
                sep=";",
                encoding="latin1",
                header=None if sem_cabecalho else "infer"  # BOOM!
            )
        else:
            raise ValueError("❌ Apenas arquivos CSV são permitidos.")


    try:
        with st.spinner("📂 Carregando planilhas..."):
            df_erp = carregar_planilha(caminho_erp)
            df_credshop = carregar_planilha(caminho_credshop, sem_cabecalho=True)  # força header=None

            with st.spinner("🔧 Iniciando limpeza e conciliação dos dados..."):
                df_erp = limpar_erp(df_erp)
                df_credshop = limpar_credshop(df_credshop)
                renomear_colunas_credshop(df_credshop)
                df_conciliado, df_erp = conciliar_credshop_erp(df_credshop, df_erp)
                df_aba_conciliados = df_conciliado[df_conciliado["Pontuação"] != 999].copy()
                df_aba_nao_conciliados = df_conciliado[df_conciliado["Pontuação"] == 999].copy()
                # Remover "aluguéis" e "estornos" da aba "Não conciliados"
                if "TIPO DE LANÇAMENTO CREDISHOP" in df_aba_nao_conciliados.columns:
                    tipo_lcto = df_aba_nao_conciliados["TIPO DE LANÇAMENTO CREDISHOP"].str.lower()
                    df_aba_nao_conciliados = df_aba_nao_conciliados[~tipo_lcto.str.contains("aluguel", na=False)]
                    df_aba_nao_conciliados = df_aba_nao_conciliados[~tipo_lcto.str.contains("estorno", na=False)]




        totais_conc = {
            "liquido": df_aba_conciliados["VALOR LIQUIDO CREDISHOP"].sum(),
            "parcela": df_aba_conciliados["VALOR BRUTO CREDISHOP"].sum(),
            "qtd": len(df_aba_conciliados)
        }
        totais_nao = {
            "liquido": df_aba_nao_conciliados["VALOR LIQUIDO CREDISHOP"].sum(),
            "parcela": df_aba_nao_conciliados["VALOR BRUTO CREDISHOP"].sum(),
            "qtd": len(df_aba_nao_conciliados)
        }

        relatorio_linhas = [
            ["RELATÓRIO DE CONCILIAÇÃO", "", ""],
            ["CONCILIADO", "", ""],
            ["- Valor Líquido Total", "", f"R$ {totais_conc['liquido']:,.2f}"],
            ["- Valor da Parcela Total", "", f"R$ {totais_conc['parcela']:,.2f}"],
            ["- Quantidade de Títulos", "", f"{totais_conc['qtd']}"],
            ["", "", ""],
            ["NÃO CONCILIADO", "", ""],
            ["- Valor Líquido Total", "", f"R$ {totais_nao['liquido']:,.2f}"],
            ["- Valor da Parcela Total", "", f"R$ {totais_nao['parcela']:,.2f}"],
            ["- Quantidade de Títulos", "", f"{totais_nao['qtd']}"]
        ]
        relatorio_df = pd.DataFrame(relatorio_linhas, columns=["Categoria", "Descrição", "VALOR BRUTO ERP"])

        # =====================================================================
        # EXCLUSÃO FINAL DAS COLUNAS (APÓS TODO O PROCESSAMENTO)
        # =====================================================================
        # Definir colunas a serem excluídas pelos nomes reais
        colunas_para_excluir = [
            "TAXA CREDISHOP",
        ]
        
        # Aplicar exclusão apenas se as colunas existirem
        for col in colunas_para_excluir:
            if col in df_aba_conciliados.columns:
                df_aba_conciliados = df_aba_conciliados.drop(columns=[col])
            if col in df_aba_nao_conciliados.columns:
                df_aba_nao_conciliados = df_aba_nao_conciliados.drop(columns=[col])
        
        # Agora gerar o Excel com as colunas já excluídas
        output_path = "Conciliação_final.xlsx"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_aba_conciliados.to_excel(writer, sheet_name="Conciliados", index=False)
            df_aba_nao_conciliados.to_excel(writer, sheet_name="Não conciliados", index=False)
            relatorio_df.to_excel(writer, sheet_name="Resumo", index=False)

            if "TIPO DE LANÇAMENTO CREDISHOP" in df_credshop.columns:
                df_credshop["TIPO DE LANÇAMENTO CREDISHOP"] = df_credshop["TIPO DE LANÇAMENTO CREDISHOP"].astype(str)

                df_aluguel = df_credshop[df_credshop["TIPO DE LANÇAMENTO CREDISHOP"].str.lower().str.contains("aluguel", na=False)]
                if not df_aluguel.empty:
                    df_aluguel.to_excel(writer, sheet_name="Aluguel", index=False)

                df_estorno = df_credshop[df_credshop["TIPO DE LANÇAMENTO CREDISHOP"].str.lower().str.contains("estorno", na=False)]
                if not df_estorno.empty:
                    df_estorno.to_excel(writer, sheet_name="Estorno", index=False)

            if "Sheet1" in writer.book.sheetnames:
                writer.book.remove(writer.book["Sheet1"])

        # === INSERE OS BLOCOS DE CHAVE ERP NA ABA RESUMO ===
        try:
            wb = load_workbook(output_path)
            ws_conciliados = wb["Conciliados"]
            ws_resumo = wb["Resumo"]

            # Identifica a coluna "CHAVE ERP" dinamicamente
            header = [cell.value for cell in ws_conciliados[1]]
            if "CHAVE ERP" in header:
                idx_chave = header.index("CHAVE ERP")
                letra_coluna = chr(65 + idx_chave)  # converte índice em letra (A=65)

                # Coleta os valores da coluna usando a letra encontrada
                col_chave = ws_conciliados[letra_coluna]
                chaves = [str(cell.value) for cell in col_chave[1:] if cell.value is not None]
                
                # Divide em blocos de 2000
                blocos = [chaves[i:i+2000] for i in range(0, len(chaves), 2000)]
                blocos_concat = [", ".join(bloco) for bloco in blocos]

                # Adiciona na aba Resumo
                start_row = ws_resumo.max_row + 2
                for i, texto in enumerate(blocos_concat, start=1):
                    ws_resumo.cell(row=start_row + i - 1, column=1, value=f"Grupo {i}")
                    ws_resumo.cell(row=start_row + i - 1, column=2, value=texto)

                wb.save(output_path)
            else:
                st.warning("Coluna 'CHAVE ERP' não encontrada na aba Conciliados")
        except Exception as e:
            st.error(f"Erro ao adicionar blocos de CHAVE ERP: {e}")

        # INTERFACE FINAL
        with st.container():
            st.header("Resultados da Conciliação")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("✅ Conciliados", 
                        f"R$ {totais_conc['liquido']:,.2f}", 
                        f"{totais_conc['qtd']} títulos")
            with col2:
                st.metric("⚠ Não Conciliados", 
                        f"R$ {totais_nao['liquido']:,.2f}", 
                        f"{totais_nao['qtd']} títulos")

            with st.expander("📊 Ver relatório completo"):
                st.dataframe(relatorio_df, hide_index=True)

        if os.path.exists(output_path):
            with open(output_path, "rb") as f:
                st.download_button(
                    label="📥 Baixar Planilha Final",
                    data=f,
                    file_name="Conciliação_final_credshop.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"❌ Erro ao carregar arquivos: {e}")
        st.stop()