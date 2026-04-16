#SANTANDER FUNCIONANDO - Ultima atualização: 16/04/2026
#____________________________________________________________________________________________________________________________________________________________________________________________________________________________
# Importação das bibliotecas necessárias:
import pandas as pd
import logging
import numpy as np
import streamlit as st
import os
import sys
from rapidfuzz import process, fuzz
from openpyxl import load_workbook
from pandas import ExcelWriter

def main():
# Configuração de logging
    logging.basicConfig(
        level=logging.DEBUG,  # ou DEBUG para mais detalhes
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler("conciliacao.log", encoding="utf-8"),  # grava em arquivo
            logging.StreamHandler()  # mostra no console
        ]
    )


    def resource_path(relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)


    # Função de carregamento
    def carregar_planilha(caminho):
        if caminho.name.endswith(".csv"):
            return pd.read_csv(caminho, sep=";", encoding="latin1", dtype={"NSU": str})
        else:
            return pd.read_excel(caminho, sheet_name="Detalhado", dtype={"NÚMERO COMPROVANTE DE VENDA (NSU)": str})
        
    # --- BARRA LATERAL ---
    # --- BARRA LATERAL ---
    with st.sidebar:
        st.markdown("# App Conciliação Bancária")
        
        # Seção de upload com tratamento de None
        st.markdown("### Carregar planilhas")
        caminho_erp = st.file_uploader("ERP (CSV)", type=["csv"], key="erp_uploader")
        caminho_santander = st.file_uploader("Santander (XLSX)", type=["xlsx"], key="santander_uploader")

    # --- ÁREA PRINCIPAL ---

    if caminho_erp is None or caminho_santander is None:
        st.subheader("Bem-vindo ao Sistema de Conciliação")
        st.markdown("""
        <div style='text-align: center; margin-bottom: 20px;'>
            <p>Este sistema realiza a conciliação automática entre:</p>
            <p>•  Santander</p>
            <p>• ERP</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.warning("⚠️ Por favor, faça upload de ambos os arquivos para iniciar a conciliação")
        
        
        st.stop()
    try:
        with st.spinner('📂 Carregando planilhas...'):
            df_erp = carregar_planilha(caminho_erp)
            df_santander = carregar_planilha(caminho_santander)
    except Exception as e:
        st.error(f"❌ Erro ao carregar arquivos: {str(e)}")
        st.stop()

    # --- Processamento
    with st.spinner('🔧 Processando dados do Santander...'):
        def limpar_santander(df):
            df = df.iloc[6:].reset_index(drop=False)
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)
            return df

    df_santander = limpar_santander(df_santander)
    df_santander = df_santander.filter(items=["EC CENTRALIZADOR", "DATA DE VENCIMENTO", "TIPO DE LANÇAMENTO", "PARCELAS", "AUTORIZAÇÃO", "NÚMERO COMPROVANTE DE VENDA (NSU)", "DATA DA VENDA", "VALOR DA PARCELA", "VALOR LÍQUIDO", "BANDEIRA / MODALIDADE"])

    #Convertendo colunas para número
    df_santander["VALOR LÍQUIDO"] = pd.to_numeric(df_santander["VALOR LÍQUIDO"], errors="coerce")
    df_santander["VALOR DA PARCELA"] = pd.to_numeric(df_santander["VALOR DA PARCELA"], errors="coerce")


    #Convertendo parcelas para números inteiros
    df_santander[["PARCELA", "TOTAL_PARCELAS"]] = df_santander["PARCELAS"].str.extract(r"(\d+)\s+de\s+(\d+)") #Agora na planilha santander, o campo parcela vem em apenas 1 celula precisando separar em colunas.
    df_santander["PARCELA"] = pd.to_numeric(df_santander["PARCELA"], errors="coerce")
    df_santander["PARCELA"] = df_santander["PARCELA"].fillna(1).astype(int) #Essa linha converte o número da parcela do tipo float para interger porém quando a venda é no débito o mesmo vem zerado. Sendo assim optou-se por preencher esse campo como valor 1, o mesmo ocorre para quantidade de parcelas
    df_santander["TOTAL_PARCELAS"] = pd.to_numeric(df_santander["TOTAL_PARCELAS"], errors="coerce")
    df_santander["TOTAL_PARCELAS"] = df_santander["TOTAL_PARCELAS"].fillna(1).astype(int)

    #Convertendo Data do pagamento e Data do lançamento para data
    df_santander["DATA DA VENDA"] = pd.to_datetime(df_santander["DATA DA VENDA"], format="%d/%m/%Y", errors="coerce")
    df_santander["DATA DE VENCIMENTO"] = pd.to_datetime(df_santander["DATA DE VENCIMENTO"], format="%d/%m/%Y", errors="coerce")

    #Separando os valores de aluguel de máquina e cancelamento dos valores da GETNET.
    df_cancelamento_venda = df_santander[df_santander["TIPO DE LANÇAMENTO"] == "Cancelamento/Chargeback"]
    df_aluguel_maquina = df_santander[df_santander["TIPO DE LANÇAMENTO"] == "Aluguel/Tarifa"]

    #Atualizando a tabela df_santander para todos os valores sem o aluguel de máquina, sem cancelamento e sem valores em branco
    df_santander = df_santander[df_santander["TIPO DE LANÇAMENTO"] != "Cancelamento/Chargeback"]
    df_santander = df_santander[df_santander["TIPO DE LANÇAMENTO"] != "Aluguel/Tarifa"]
    df_santander = df_santander[df_santander["TIPO DE LANÇAMENTO"] != "Pagamento Realizado"]
    df_santander = df_santander[df_santander["TIPO DE LANÇAMENTO"] != "Saldo Anterior"]
    df_santander = df_santander[df_santander["TIPO DE LANÇAMENTO"].notna()]


    #Totalizadores
    valor_total_bruto = df_santander["VALOR DA PARCELA"].sum()
    quantidade_titulos_santander = df_santander["VALOR DA PARCELA"].count()
    valor_total_liquido =  df_santander["VALOR LÍQUIDO"].sum()
    valor_aluguel_maquina = df_aluguel_maquina["VALOR LÍQUIDO"].sum()
    valor_cancelamento_venda = df_cancelamento_venda["VALOR LÍQUIDO"].sum()
    quantidade_titulos_cancelados = df_cancelamento_venda["VALOR LÍQUIDO"].count()
    valor_recebido_conta = valor_total_liquido - abs(valor_aluguel_maquina) - abs(valor_cancelamento_venda)


    #Selecionando as colunas desejadas
    with st.spinner('🛠️ Processando dados do ERP...'):
        df_erp = df_erp.filter(items=["1o. Agrupamento", "Chave", "Numero", "NSU", "Autorização", "Emissão", "Correção", "Valor", "Vr Corrigido", "Pessoa do Título"])
    #Convertendo colunas para os tipos corretos
    #Convertendo colunas para número
    df_erp["Valor"] = df_erp["Valor"].str.replace(",", ".", regex=True)
    df_erp["Valor"] = pd.to_numeric(df_erp["Valor"], errors="coerce")
    df_erp["Vr Corrigido"] = df_erp["Vr Corrigido"].str.replace(",", ".", regex=True)
    df_erp["Vr Corrigido"] = pd.to_numeric(df_erp["Vr Corrigido"], errors="coerce")
    #Convertendo colunas para data
    df_erp["Emissão"] = pd.to_datetime(df_erp["Emissão"], format="%d/%m/%Y", errors="coerce")
    df_erp["Correção"] = pd.to_datetime(df_erp["Correção"], format="%d/%m/%Y", errors="coerce")
    #Transformando o campo Numero em Parcela e Total de Parcelas
    # Criar as novas colunas extraindo os valores corretos da coluna "Numero"
    df_erp["chcriacao"] = df_erp["Numero"].str.split("-").str[0]  # Antes do "-"
    df_erp["Parcela"] = df_erp["Numero"].str.split("-").str[1].str.split("/").str[0]  # Entre "-" e "/"
    df_erp["Total_Parcelas"] = df_erp["Numero"].str.split("/").str[1]  # Após "/"

    # Converter as colunas de parcela para inteiro
    df_erp["Parcela"] = pd.to_numeric(df_erp["Parcela"], errors="coerce").fillna(1).astype(int)
    df_erp["Total_Parcelas"] = pd.to_numeric(df_erp["Total_Parcelas"], errors="coerce").fillna(1).astype(int)
    df_erp = df_erp.filter(items=["1o. Agrupamento", "Chave", "chcriacao", "Parcela", "Total_Parcelas", "NSU", "Autorização", "Emissão", "Correção", "Valor", "Vr Corrigido", "Pessoa do Título"])

    #Selecionando apenas os títulos da industria
    df_erp_loja = df_erp[~df_erp["1o. Agrupamento"].isin(["LE SFR Indústria Ltda", "LE Protendidos"])].copy()



    #Funções Conciliar por valor e data e conciliando buscando autorizações parecidas
    with st.spinner('🔍 Realizando conciliação...'):
        def conciliar_por_data_e_valores(row, df_erp_base):

        # 1️ Filtra por datas com até 5 dias de diferença
            data_diferenca = (df_erp_base["Emissão"] - row["DATA DA VENDA"]).abs().dt.days


            candidatos = df_erp_base[data_diferenca <= 5]


        # 2️ Filtra por valor, parcela e total de parcelas
            candidatos = candidatos[
                ((candidatos["Valor"] - row["VALOR DA PARCELA"]).abs() <= 0.20) &
                (candidatos["Parcela"] == row["PARCELA"]) &
                (candidatos["Total_Parcelas"] == row["TOTAL_PARCELAS"])
        ]

            if not candidatos.empty:
                linha = candidatos.iloc[0]

                return pd.Series([
                    linha["Autorização"],
                    linha["Chave"],
                    linha["Valor"],
                    "Conciliado por Data e Valores",
                10
            ])                      
            return pd.Series([None, None, None, "Não Conciliado", 99])

    with st.spinner('🔎 Realizando conciliação...'):
        def encontrar_melhor_correspondencia_com_pontuacao(row, df_origem, coluna_erp):
            correspondencias = process.extract(
                str(row["AUTORIZAÇÃO"]),
                df_origem[coluna_erp].astype(str),
                scorer=fuzz.ratio,
                limit=10
            )

            correspondencias_validas = [(texto, score, idx) for texto, score, idx in correspondencias if score >= 80]



            if not correspondencias_validas:
                return pd.Series([None, None, None, "Não Conciliado", 99])

            melhor_resultado = None
            menor_pontuacao = float("inf")

            for melhor_correspondencia, melhor_pontuacao, _ in correspondencias_validas:
                filtro = df_origem[df_origem[coluna_erp] == melhor_correspondencia]

                if filtro.empty:                
                    continue

                #  Itera sobre todas as linhas com o mesmo valor
                for _, linha_correspondente in filtro.iterrows():
                    valor_erp = linha_correspondente["Valor"]
                    data_erp = linha_correspondente["Emissão"]
                    parcela_erp = linha_correspondente["Parcela"]
                    total_parcelas_erp = linha_correspondente["Total_Parcelas"]

                    status = ["Conciliado"]
                    pontuacao = 0

                    if abs(row["VALOR DA PARCELA"] - valor_erp) > 0.10:
                        status.append("Divergência de Valor")
                        pontuacao += 15

                    if abs((row["DATA DA VENDA"] - data_erp).days) > 1:
                        status.append("Divergência de Data")
                        pontuacao += 5

                    if row["PARCELA"] != parcela_erp:
                        status.append("Divergência de Parcela")
                        pontuacao += 10

                    if row["TOTAL_PARCELAS"] != total_parcelas_erp:
                        status.append("Divergência de Total de Parcelas")
                        pontuacao += 15


                    if pontuacao < menor_pontuacao:
                        menor_pontuacao = pontuacao
                        melhor_resultado = (
                            linha_correspondente[coluna_erp],
                            linha_correspondente["Chave"],
                            valor_erp,
                            " com ".join(status) if len(status) > 1 else status[0],
                            pontuacao
                        )

            if melhor_resultado:

                return pd.Series(melhor_resultado)
            else:

                return pd.Series([None, None, None, "Não Conciliado", 99])
            
        def encontrar_melhor_correspondencia_com_pontuacao_nsu(row, df_origem):
            correspondencias = process.extract(
                str(row["NÚMERO COMPROVANTE DE VENDA (NSU)"]),
                df_origem["NSU"].astype(str),
                scorer=fuzz.ratio,
                limit=10
            )

            correspondencias_validas = [(texto, score, idx) for texto, score, idx in correspondencias if score >= 80]

            print(f"\n Buscando correspondência para: {row['NÚMERO COMPROVANTE DE VENDA (NSU)']}")
            print("Correspondências válidas (score >= 80):", correspondencias_validas)

            if not correspondencias_validas:
                return pd.Series([None, None, None, "Não Conciliado", 99])

            melhor_resultado = None
            menor_pontuacao = float("inf")

            for melhor_correspondencia, melhor_pontuacao, _ in correspondencias_validas:
                filtro = df_origem[df_origem["NSU"] == melhor_correspondencia]

                if filtro.empty:
                    print(f"⚠ Correspondência '{melhor_correspondencia}' não encontrada no DataFrame.")
                    continue

                #  Itera sobre todas as linhas com o mesmo valor
                for _, linha_correspondente in filtro.iterrows():
                    valor_erp = linha_correspondente["Valor"]
                    data_erp = linha_correspondente["Emissão"]
                    parcela_erp = linha_correspondente["Parcela"]
                    total_parcelas_erp = linha_correspondente["Total_Parcelas"]

                    status = ["Conciliado"]
                    pontuacao = 0

                    if abs(row["VALOR DA PARCELA"] - valor_erp) > 0.10:
                        status.append("Divergência de Valor")
                        pontuacao += 15

                    if abs((row["DATA DA VENDA"] - data_erp).days) > 1:
                        status.append("Divergência de Data")
                        pontuacao += 5

                    if row["PARCELA"] != parcela_erp:
                        status.append("Divergência de Parcela")
                        pontuacao += 10

                    if row["TOTAL_PARCELAS"] != total_parcelas_erp:
                        status.append("Divergência de Total de Parcelas")
                        pontuacao += 15

                    if pontuacao < menor_pontuacao:
                        menor_pontuacao = pontuacao
                        melhor_resultado = (
                            linha_correspondente["NSU"],
                            linha_correspondente["Chave"],
                            valor_erp,
                            " e ".join(status) if len(status) > 1 else status[0],
                            pontuacao
                        )

            if melhor_resultado:
                print(" Melhor resultado escolhido:", melhor_resultado)
                return pd.Series(melhor_resultado)
            else:
                print(" Nenhuma correspondência com pontuação aceitável.")
                return pd.Series([None, None, None, "Não Conciliado", 99])

        def selecionar_melhor_por_pontuacao_com_autorizacao_e_nsu(row, df_erp_base, tolerancia_dias=5, tolerancia_valor=0.20, incluir_detalhes=False):
            candidatos = df_erp_base[
                (df_erp_base["Emissão"] - row["DATA DA VENDA"]).abs().dt.days <= tolerancia_dias
            ]

            candidatos = candidatos[
                (candidatos["Valor"] - row["VALOR DA PARCELA"]).abs() <= tolerancia_valor
            ]

            candidatos = candidatos[
                (candidatos["Parcela"] == row["PARCELA"]) &
                (candidatos["Total_Parcelas"] == row["TOTAL_PARCELAS"])
            ]

            if candidatos.empty:
                if incluir_detalhes:
                    return pd.Series([None, None, None, None, None, None, "Não Conciliado", 999])
                else:
                    return pd.Series([None, None, None, None, "Não Conciliado", 999])

            melhor_resultado = None
            menor_pontuacao = float("inf")

            for _, linha in candidatos.iterrows():
                if pd.isna(linha["Emissão"]) or pd.isna(row["DATA DA VENDA"]):
                    dias_dif = 999
                else:
                    dias_dif = abs((linha["Emissão"] - row["DATA DA VENDA"]).days)
                    
                valor_dif = abs(linha["Valor"] - row["VALOR DA PARCELA"])

                aut_sant = str(row["AUTORIZAÇÃO"]).strip()
                aut_erp = str(linha["Autorização"]).strip()
                nsu_sant = str(row["NÚMERO COMPROVANTE DE VENDA (NSU)"]).strip()
                nsu_erp = str(linha["NSU"]).strip()

                if aut_sant == aut_erp or nsu_sant == nsu_erp:
                    sim_autorizacao = 100
                    sim_nsu = 100
                else:
                    sim_autorizacao = fuzz.ratio(aut_sant, aut_erp)
                    sim_nsu = fuzz.ratio(nsu_sant, nsu_erp)

                pontuacao = dias_dif * 100 + valor_dif * 100 + (200 - (sim_autorizacao + sim_nsu))
                if "Pessoa do Título" in linha and linha["Pessoa do Título"] != "Getnet Adquirencia E Servicos Para Meios de Pagamento S.a.":
                    pontuacao += 101

                if pontuacao < menor_pontuacao:
                    menor_pontuacao = pontuacao
                    
                    # --- LÓGICA DO STATUS DINÂMICO CORRIGIDA ---
                    status_lista = ["Conciliado"]
                    
                    if valor_dif > 0.10:
                        status_lista.append("Divergência de Valor")
                    
                    if dias_dif > 1 and dias_dif != 999:
                        status_lista.append("Divergência de Data")
                    elif dias_dif == 999:
                        status_lista.append("Data Ausente")

                    # NOVA VERIFICAÇÃO: Compara se os campos são diferentes de fato
                    if (aut_sant != aut_erp) and (nsu_sant != nsu_erp):
                        status_lista.append("Divergência de NSU/Autorização")
                    elif aut_sant != aut_erp:
                        status_lista.append("Divergência de Autorização")
                    elif nsu_sant != nsu_erp:
                        status_lista.append("Divergência de NSU")
                        
                    status_final = " e ".join(status_lista) if len(status_lista) > 1 else status_lista[0]

                    melhor_resultado = (
                        linha["Autorização"],
                        linha["NSU"],
                        linha["Chave"],
                        linha["Valor"],
                        dias_dif,
                        valor_dif,
                        status_final, 
                        round(pontuacao, 2)
                    )

            if melhor_resultado:
                if incluir_detalhes:
                    return pd.Series(melhor_resultado)
                else:
                    return pd.Series(melhor_resultado[:4] + melhor_resultado[-2:])
            else:
                if incluir_detalhes:
                    return pd.Series([None, None, None, None, None, None, "Não Conciliado", 999])
                else:
                    return pd.Series([None, None, None, None, "Não Conciliado", 999])
                
        def marcar_duplicados_com_pior_score(df, chave_col="Chave ERP", status_col="Status", pontuacao_col="Pontuação"):
            # 1️ Filtra linhas com chaves duplicadas
            duplicadas = df[df.duplicated(subset=[chave_col], keep=False)].copy()

            if duplicadas.empty:
                return df


            # 2️ Ordena pela pontuação crescente (menor pontuação é a melhor)
            duplicadas_sorted = duplicadas.sort_values(pontuacao_col, ascending=True)

            # 3️ Marca como duplicado todas as duplicatas exceto a com menor pontuação
            duplicadas_marcadas = duplicadas_sorted.duplicated(subset=[chave_col], keep="first")

            # 4️ Atualiza status e pontuação das duplicadas com pior score
            df.loc[duplicadas_sorted[duplicadas_marcadas].index, status_col] = "Valor Duplicado Menor Score"
            df.loc[duplicadas_sorted[duplicadas_marcadas].index, pontuacao_col] = 998


            return df


        #Remover da Planilha Santander os Títulos que foram cancelados
        # 1️ Criar coluna auxiliar com valor absoluto da parcela
        df_santander["VALOR_ABS"] = df_santander["VALOR DA PARCELA"].abs()
        df_cancelamento_venda["VALOR_ABS"] = df_cancelamento_venda["VALOR DA PARCELA"].abs()

        # 2️ Criar chave composta: AUTORIZAÇÃO + VALOR_ABS
        df_santander["CHAVE_CONCILIACAO"] = df_santander["AUTORIZAÇÃO"].astype(str) + "_" + df_santander["VALOR_ABS"].astype(str)
        df_cancelamento_venda["CHAVE_CONCILIACAO"] = df_cancelamento_venda["AUTORIZAÇÃO"].astype(str) + "_" + df_cancelamento_venda["VALOR_ABS"].astype(str)

        # 3️ Verificar chaves em comum
        chaves_comuns = set(df_santander["CHAVE_CONCILIACAO"]) & set(df_cancelamento_venda["CHAVE_CONCILIACAO"])

        # 4️ Filtrar as linhas da df_santander que estão na lista de cancelamentos
        filtro_cancelados = df_santander["CHAVE_CONCILIACAO"].isin(df_cancelamento_venda["CHAVE_CONCILIACAO"])

        # 5️ Copiar essas linhas
        df_cancelados_encontrados = df_santander[filtro_cancelados].copy()

        # 6️ Adicionar ao df_cancelamento_venda
        df_cancelamento_venda = pd.concat([df_cancelamento_venda, df_cancelados_encontrados], ignore_index=True)

        # 7️ Remover da df_santander
        df_santander = df_santander[~filtro_cancelados].copy()

        # 8️ Resultado final

        df_primeira_conciliacao = df_santander
        df_segunda_conciliacao = df_primeira_conciliacao.filter(items=["EC CENTRALIZADOR", "DATA DE VENCIMENTO", "TIPO DE LANÇAMENTO", "PARCELAS", "AUTORIZAÇÃO", "NÚMERO COMPROVANTE DE VENDA (NSU)", "DATA DA VENDA","VALOR DA PARCELA", "VALOR LÍQUIDO", "PARCELA", "TOTAL_PARCELAS"])
        progress_bar = st.progress(0, text="🔄 Conciliando registros...")
        resultados = []

        total = len(df_segunda_conciliacao)

        for i, (_, row) in enumerate(df_segunda_conciliacao.iterrows()):
            resultado = selecionar_melhor_por_pontuacao_com_autorizacao_e_nsu(row, df_erp)
            resultados.append(resultado)
            
            progresso = (i + 1) / total
            progress_bar.progress(progresso, text=f"🔄 Conciliando ({i + 1}/{total}) registros...")

        # Coloca os resultados de volta no DataFrame
        df_segunda_conciliacao[["Autorização ERP", "NSU ERP", "Chave ERP", "Valor ERP", "Status", "Pontuação"]] = pd.DataFrame(resultados, index=df_segunda_conciliacao.index)


        df_terceira_conciliacao = df_segunda_conciliacao[df_segunda_conciliacao["Pontuação"] == 999].copy()
        df_segunda_conciliacao = df_segunda_conciliacao[df_segunda_conciliacao["Pontuação"] != 999].copy()
        df_segunda_conciliacao = marcar_duplicados_com_pior_score(df_segunda_conciliacao)
        duplicados = df_segunda_conciliacao[df_segunda_conciliacao["Pontuação"] == 998].copy()
        df_segunda_conciliacao = df_segunda_conciliacao[df_segunda_conciliacao["Pontuação"] != 998].copy()
        df_terceira_conciliacao = pd.concat([df_terceira_conciliacao, duplicados], ignore_index=True)


        df_conciliado = df_segunda_conciliacao
        df_nao_conciliado = df_terceira_conciliacao


        #Marcar na planilha ERP o que já foi usado na conciliação para não ser usado novamente.

        def marcar_e_filtrar_chaves_utilizadas(df_erp, df_conciliado):
            """
            """
            # Normaliza os valores para garantir comparação precisa
            # - Trata como string (preserva valores não-numéricos)
            # - Remove espaços em branco e sufixo '.0' gerado por conversões
            # - Converte representações 'nan' / 'None' / '' para pd.NA
            df_erp["Chave"] = (
                df_erp["Chave"].astype(str)
                .str.strip()
                .str.replace(r"\.0$", "", regex=True)
                .replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})
                .astype("string")
            )

            df_conciliado["Chave ERP"] = (
                df_conciliado["Chave ERP"].astype(str)
                .str.strip()
                .str.replace(r"\.0$", "", regex=True)
                .replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})
                .astype("string")
            )

            # Coleta as chaves que já foram utilizadas (ignorando NA)
            chaves_utilizadas = df_conciliado["Chave ERP"].dropna().unique()

            # Marca no df_erp quais foram utilizadas
            df_erp["Usada"] = df_erp["Chave"].isin(chaves_utilizadas)

            # Filtra as que ainda estão disponíveis para nova conciliação
            df_erp_disponivel = df_erp[~df_erp["Usada"]].copy()

            return df_erp, df_erp_disponivel

        df_erp, df_erp_disponivel = marcar_e_filtrar_chaves_utilizadas(df_erp, df_conciliado)

        df_nao_conciliado[["Autorização ERP", "NSU ERP", "Chave ERP", "Valor ERP", "DIF_DIAS", "DIF_VALOR", "Status", "Pontuação"]] = df_nao_conciliado.apply(
            lambda row: selecionar_melhor_por_pontuacao_com_autorizacao_e_nsu(row, df_erp_disponivel, 30, 100000.00, True),
            axis=1
        )


    # Função para gerar o relatório formatado como DataFrame
    with st.spinner('📊 Gerando relatório final...'):
        def gerar_relatorio_df_formatado(df_conciliado, df_nao_conciliado, df_cancelamento_venda, valor_aluguel_maquina):
            # Calcula os totais diretamente dos DataFrames originais
            totais = {
                'conciliado': {
                    'liquido': df_conciliado["VALOR LÍQUIDO"].sum(),
                    'parcela': df_conciliado["VALOR DA PARCELA"].sum(),
                    'qtd': len(df_conciliado)
                },
                'nao_conciliado': {
                    'liquido': df_nao_conciliado["VALOR LÍQUIDO"].sum(),
                    'parcela': df_nao_conciliado["VALOR DA PARCELA"].sum(),
                    'qtd': len(df_nao_conciliado)
                },
                'cancelado': {
                    'liquido': df_cancelamento_venda["VALOR LÍQUIDO"].sum(),
                    'parcela': df_cancelamento_venda["VALOR DA PARCELA"].sum(),
                    'qtd': len(df_cancelamento_venda)
                },
                'aluguel': valor_aluguel_maquina,
                'total_banco': df_conciliado["VALOR LÍQUIDO"].sum() + 
                            df_nao_conciliado["VALOR LÍQUIDO"].sum() + 
                            df_cancelamento_venda["VALOR LÍQUIDO"].sum() + 
                            valor_aluguel_maquina
            }

            # Constroi a estrutura do relatório
            relatorio_dados = [
                ["RELATÓRIO DE CONCILIAÇÃO", "", ""],
                ["CONCILIADO", "", ""],
                ["- Valor Líquido Total", "", f"R$ {totais['conciliado']['liquido']:,.2f}"],
                ["- Valor da Parcela Total", "", f"R$ {totais['conciliado']['parcela']:,.2f}"],
                ["- Quantidade de Títulos", "", f"{totais['conciliado']['qtd']}"],
                ["", "", ""],
                ["NÃO CONCILIADO", "", ""],
                ["- Valor Líquido Total", "", f"R$ {totais['nao_conciliado']['liquido']:,.2f}"],
                ["- Valor da Parcela Total", "", f"R$ {totais['nao_conciliado']['parcela']:,.2f}"],
                ["- Quantidade de Títulos", "", f"{totais['nao_conciliado']['qtd']}"],
                ["", "", ""],
                ["CANCELAMENTO DE VENDA", "", ""],
                ["- Valor Líquido Total", "", f"R$ {totais['cancelado']['liquido']:,.2f}"],
                ["- Valor da Parcela Total", "", f"R$ {totais['cancelado']['parcela']:,.2f}"],
                ["- Quantidade de Títulos", "", f"{totais['cancelado']['qtd']}"],
                ["", "", ""],
                ["OUTROS", "", ""],
                ["- Valor total de aluguel de máquineta", "", f"R$ {totais['aluguel']:,.2f}"],
                ["- Valor Total no Banco", "", f"R$ {totais['total_banco']:,.2f}"]
            ]

            return pd.DataFrame(relatorio_dados, columns=["Categoria", "Descrição", "Valor"])
        # --- Exibição de Resultados no Streamlit ---
        st.header("Resultados da Conciliação")

        # Gera o relatório

        relatorio_df = gerar_relatorio_df_formatado(
            df_conciliado, 
            df_nao_conciliado, 
            df_cancelamento_venda, 
            valor_aluguel_maquina
        )

        # Exibe as métricas principais (usando valores diretos, não do DataFrame)

        with st.container():
            st.subheader("Resumo Financeiro")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("✅ Conciliados", 
                        f"R$ {df_conciliado['VALOR LÍQUIDO'].sum():,.2f}", 
                        f"{len(df_conciliado)} títulos")
            with col2:
                st.metric("⚠ Não Conciliados", 
                        f"R$ {df_nao_conciliado['VALOR LÍQUIDO'].sum():,.2f}", 
                        f"{len(df_nao_conciliado)} títulos")
            with col3:
                st.metric("❌ Cancelados", 
                        f"R$ {df_cancelamento_venda['VALOR LÍQUIDO'].sum():,.2f}", 
                        f"{len(df_cancelamento_venda)} títulos")

            # Exibe a tabela completa 
            with st.expander("📊 Ver relatório completo"):
                st.dataframe(relatorio_df, hide_index=True)

# --- 1. DEFINIÇÃO DA ORDEM DAS COLUNAS (O que o usuário vai ver) ---
        # Organizamos aqui: Primeiro ERP, depois Getnet, depois Resultados
        colunas_ordenadas = [
            # Bloco ERP
            "Cliente ERP", "Chave ERP", "Valor ERP", "Valor Bruto ERP", "NSU ERP", "Autorização ERP",
            
            # Bloco Getnet
            "Data Venda Getnet", "Valor Getnet", "Valor Líquido Getnet", 
            "NSU Getnet", "Autorização Getnet", "Data Vencimento Getnet", 
            "Parcela Atual Getnet", "Total Parcelas Getnet", "Bandeira Getnet",
            
            # Bloco de Resultados/Pontuação
            "Status da Conciliação", "Diferença Valor (Getnet x ERP)", 
            "Diferença Dias (Getnet x ERP)", "Score de Confiança"
        ]

        # Dicionário de Renomeação (Mapeia o nome técnico para o amigável)
        nomes_amigaveis = {
            "Pessoa do Título": "Cliente ERP",
            "Chave ERP": "Chave ERP",
            "Valor ERP": "Valor ERP",
            "Valor bruto": "Valor Bruto ERP",
            "NSU ERP": "NSU ERP",
            "Autorização ERP": "Autorização ERP",
            "DATA DA VENDA": "Data Venda Getnet",
            "VALOR DA PARCELA": "Valor Getnet",
            "VALOR LÍQUIDO": "Valor Líquido Getnet",
            "NÚMERO COMPROVANTE DE VENDA (NSU)": "NSU Getnet",
            "AUTORIZAÇÃO": "Autorização Getnet",
            "DATA DE VENCIMENTO": "Data Vencimento Getnet",
            "PARCELA": "Parcela Atual Getnet",
            "TOTAL_PARCELAS": "Total Parcelas Getnet",
            "BANDEIRA / MODALIDADE": "Bandeira Getnet",
            "Status": "Status da Conciliação",
            "DIF_VALOR": "Diferença Valor (Getnet x ERP)",
            "DIF_DIAS": "Diferença Dias (Getnet x ERP)",
            "Pontuação": "Score de Confiança"
        }

        colunas_para_excluir = ["TIPO DE LANÇAMENTO", "Tipo Lançamento Getnet", "Chave", "Valor Getnet", "EC Getnet", "EC CENTRALIZADOR"]
        
        
        
    output_path = "Conciliação_final.xlsx"

    # ==========================================
    # 1) VALIDAÇÕES EXPLÍCITAS ANTES DE EXPORTAR
    # ==========================================
    try:
        logging.info("=== INICIANDO DEBUG DE EXPORTAÇÃO ===")
        logging.info(f"df_conciliado shape: {df_conciliado.shape}")
        logging.info(f"df_nao_conciliado shape: {df_nao_conciliado.shape}")
        logging.info(f"df_erp shape: {df_erp.shape}")
        logging.info(f"df_cancelamento_venda shape: {df_cancelamento_venda.shape}")
        logging.info(f"df_aluguel_maquina shape: {df_aluguel_maquina.shape}")
        logging.info(f"relatorio_df shape: {relatorio_df.shape}")

        # Colunas mínimas esperadas antes do merge
        colunas_obrigatorias_conc = ["Chave ERP", "Status", "Pontuação"]
        colunas_obrigatorias_nconc = ["Chave ERP", "Status", "Pontuação"]
        colunas_obrigatorias_erp = ["Chave", "Valor", "Pessoa do Título"]

        faltando_conc = [c for c in colunas_obrigatorias_conc if c not in df_conciliado.columns]
        faltando_nconc = [c for c in colunas_obrigatorias_nconc if c not in df_nao_conciliado.columns]
        faltando_erp = [c for c in colunas_obrigatorias_erp if c not in df_erp.columns]

        if faltando_conc:
            raise KeyError(f"df_conciliado está sem colunas obrigatórias: {faltando_conc}")

        if faltando_nconc:
            raise KeyError(f"df_nao_conciliado está sem colunas obrigatórias: {faltando_nconc}")

        if faltando_erp:
            raise KeyError(f"df_erp está sem colunas obrigatórias: {faltando_erp}")

        # ==========================================
        # 2) PREPARAÇÃO DOS DATAFRAMES
        # ==========================================
        logging.info("Preparando df_conc_final...")
        df_conc_final = df_conciliado.merge(
            df_erp[["Chave", "Valor", "Pessoa do Título"]],
            left_on="Chave ERP",
            right_on="Chave",
            how="left"
        ).rename(columns={"Valor": "Valor bruto"})

        logging.info(f"df_conc_final shape: {df_conc_final.shape}")
        logging.info(f"df_conc_final columns: {list(df_conc_final.columns)}")

        logging.info("Preparando df_nconc_final...")
        df_nconc_final = df_nao_conciliado.merge(
            df_erp[["Chave", "Valor", "Pessoa do Título"]],
            left_on="Chave ERP",
            right_on="Chave",
            how="left"
        ).rename(columns={"Valor": "Valor bruto"})

        logging.info(f"df_nconc_final shape: {df_nconc_final.shape}")
        logging.info(f"df_nconc_final columns: {list(df_nconc_final.columns)}")

        logging.info("Aplicando drop/rename/reordenação...")
        df_conc_exp = df_conc_final.drop(columns=colunas_para_excluir, errors="ignore").rename(columns=nomes_amigaveis)
        df_nconc_exp = df_nconc_final.drop(columns=colunas_para_excluir, errors="ignore").rename(columns=nomes_amigaveis)

        cols_presentes_conc = [c for c in colunas_ordenadas if c in df_conc_exp.columns]
        cols_presentes_nconc = [c for c in colunas_ordenadas if c in df_nconc_exp.columns]

        logging.info(f"cols_presentes_conc: {cols_presentes_conc}")
        logging.info(f"cols_presentes_nconc: {cols_presentes_nconc}")

        # Se isso aqui falhar, você achou a origem real do problema
        if not cols_presentes_conc:
            raise ValueError(
                "Nenhuma coluna esperada apareceu em df_conc_exp após rename(). "
                "Verifique nomes em nomes_amigaveis e colunas reais do DataFrame."
            )

        if not cols_presentes_nconc:
            raise ValueError(
                "Nenhuma coluna esperada apareceu em df_nconc_exp após rename(). "
                "Verifique nomes em nomes_amigaveis e colunas reais do DataFrame."
            )

    except Exception as e:
        logging.exception("Falha na preparação dos dados antes da exportação")
        st.error(f"❌ Erro na preparação dos dados: {e}")
        raise

    # ==========================================
    # 3) ESCRITA DO EXCEL
    # ==========================================
    try:
        with st.spinner("Organizando colunas (ERP > Getnet > Pontuação)..."):
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                logging.info("Escrevendo aba Conciliados...")
                df_conc_exp.loc[:, cols_presentes_conc].to_excel(
                    writer,
                    sheet_name="Conciliados",
                    index=False
                )

                logging.info("Escrevendo aba Não conciliados...")
                df_nconc_exp.loc[:, cols_presentes_nconc].to_excel(
                    writer,
                    sheet_name="Não conciliados",
                    index=False
                )

                logging.info("Escrevendo aba Cancelamentos...")
                df_cancelamento_venda.to_excel(
                    writer,
                    sheet_name="Cancelamentos",
                    index=False
                )

                logging.info("Escrevendo aba Aluguel e Tarifas...")
                df_aluguel_maquina.to_excel(
                    writer,
                    sheet_name="Aluguel e Tarifas",
                    index=False
                )

                logging.info("Escrevendo aba Resumo...")
                relatorio_df.to_excel(
                    writer,
                    sheet_name="Resumo",
                    index=False
                )

        logging.info("Arquivo Excel gravado com sucesso. Abrindo com openpyxl...")

        wb = load_workbook(output_path)

        logging.info(f"Aba(s) encontradas no workbook: {wb.sheetnames}")

        if "Conciliados" not in wb.sheetnames:
            raise KeyError("A aba 'Conciliados' não existe no arquivo final.")

        if "Resumo" not in wb.sheetnames:
            raise KeyError("A aba 'Resumo' não existe no arquivo final.")

        ws_conciliados = wb["Conciliados"]
        ws_resumo = wb["Resumo"]

        header = [cell.value for cell in ws_conciliados[1]]
        logging.info(f"Header da aba Conciliados: {header}")

        if "Chave ERP" in header:
            idx_chave = header.index("Chave ERP")
            letra_coluna = ws_conciliados.cell(row=1, column=idx_chave + 1).column_letter

            chaves = [
                str(cell.value)
                for cell in ws_conciliados[letra_coluna][1:]
                if cell.value is not None
            ]

            logging.info(f"Quantidade de chaves ERP encontradas: {len(chaves)}")

            blocos = [chaves[i:i + 2000] for i in range(0, len(chaves), 2000)]
            blocos_concat = [", ".join(bloco) for bloco in blocos]

            start_row = ws_resumo.max_row + 2
            for i, texto in enumerate(blocos_concat, start=1):
                ws_resumo.cell(row=start_row + i - 1, column=1, value=f"Grupo {i}")
                ws_resumo.cell(row=start_row + i - 1, column=2, value=texto)

        wb.save(output_path)

        st.download_button(
            label="📥 Baixar Planilha Final",
            data=open(output_path, "rb"),
            file_name="Conciliacao_Final_Getnet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logging.exception("Falha na escrita/posprocessamento do Excel")
        st.error(f"❌ Erro real encontrado: {e}")
        raise