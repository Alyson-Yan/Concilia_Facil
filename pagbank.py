#===========================================================================
#IMPORTA AS BIBLIOTECAS NECESSÁRIAS
#===========================================================================
import os
import re  
import logging
import numpy as np
import pandas as pd
import streamlit as st
from io import BytesIO
from rapidfuzz import fuzz
from openpyxl.styles import Font


# Configuração de logging
logging.basicConfig(
    level=logging.DEBUG, 
    format="%(asctime)s - %(levelname)s - %(message)s", 
    handlers=[
        logging.FileHandler("conciliacao.log", encoding="utf-8"),  
        logging.StreamHandler()  # Exibe logs no console
    ]
)

#===========================================================================
# PADRONIZAÇÃO DO DATAFRAME DO PAGBANK
#===========================================================================
def padronizar_pagbank(df):  # Função para padronizar o DataFrame do PagBank

    # renomeia colunas para padronizar com sufixo Pagbank
    df = df.rename(columns={
        "Código da Transação": "Código da Transação Pagbank",
        "Data da Transação": "Data da Transação Pagbank",
        "Data prevista de liberação": "Data prevista de liberação Pagbank",
        "Bandeira": "Bandeira Pagbank",
        "Forma de Pagamento": "Forma de Pagamento Pagbank",
        "Parcela": "Parcela Pagbank",
        "Valor Bruto": "Valor Bruto Pagbank",
        "Valor Taxa": "Valor Taxa Pagbank",
        "Valor Líquido": "Valor Líquido Pagbank",
        "Código NSU": "Código da Venda Pagbank", 
        "Código de Autorização": "Código de Autorização Pagbank",
        "Identificação da Maquininha": "Identificação da Maquininha Pagbank",
        "Código da Venda": "Código NSU Pagbank",    })  # Renomeia colunas para consistência

    # converte data da transação para datetime sem hora
    df["Data da Transação Pagbank"] = pd.to_datetime(
        df["Data da Transação Pagbank"], dayfirst=True, errors="coerce"
    ).dt.normalize()  # Converte para datetime, assume dia primeiro, remove hora

    # função para limpar e converter valores monetários para float
    def converter_valor(col):  # Função interna para converter strings monetárias em float
        return (
            col.astype(str)
            .str.replace("R$", "", regex=False)  # Remove símbolo R$
            .str.replace(".", "", regex=False)  # Remove ponto de milhar
            .str.replace(",", ".", regex=False)  # Substitui vírgula por ponto
            .astype(float)  # Converte para float
        )

    # aplica conversão nos valores financeiros
    df["Valor Bruto Pagbank"] = converter_valor(df["Valor Bruto Pagbank"])  # Converte valor bruto
    df["Valor Taxa Pagbank"] = converter_valor(df["Valor Taxa Pagbank"])  # Converte taxa
    df["Valor Líquido Pagbank"] = converter_valor(df["Valor Líquido Pagbank"])  # Converte valor líquido

    # garante que identificadores sejam tratados como texto
    df["Código da Transação Pagbank"] = df["Código da Transação Pagbank"].astype(str)  # Converte para string
    df["Código NSU Pagbank"] = df["Código NSU Pagbank"].astype(str)  # Converte para string
    df["Código de Autorização Pagbank"] = df["Código de Autorização Pagbank"].astype(str)  # Converte para string
    df["Código da Venda Pagbank"] = df["Código da Venda Pagbank"].astype(str)  # Converte para string

    def extrair_parcelas_pagbank(valor):  # Função para extrair parcela atual e total de parcelas
        if pd.isna(valor):  # Se valor é NaN, assume 1/1
            return 1, 1

        texto = str(valor).lower().strip()  # Converte para string minúscula e remove espaços

        # padrão tipo: "1/3"
        match_barra = re.search(r"(\d+)\s*/\s*(\d+)", texto)  # Busca padrão com barra

        if match_barra:
            return int(match_barra.group(1)), int(match_barra.group(2))  # Retorna parcela e total

        # padrão tipo: "3x" (sem saber a parcela atual)
        match_x = re.search(r"(\d+)\s*x", texto)  # Busca padrão com x

        if match_x:
            total = int(match_x.group(1))  # Total de parcelas
            return 1, total  # Assume primeira parcela

        return 1, 1  # Padrão se não encontrar


    # cria as duas colunas padrão ERP-like
    df[["Parcela Pagbank", "Total Parcelas Pagbank"]] = df["Parcela Pagbank"] \
        .apply(lambda x: pd.Series(extrair_parcelas_pagbank(x)))  # Aplica função e cria colunas

    # lista de colunas a excluir
    df_excluir = [
        "Nome Cliente",
        "E-mail Cliente",
        "Código da Transação Pagbank",
        "Documento",
        "Status",
        "Valor Repasse",
        "Estabelecimento",
        "Código Referência",
        "Nome Comprador",
        "E-mail Comprador",
        "Código TX ID (PIX)",
        "ID Split",
        "Número do Cartão",
        "Data do cancelamento",
    ]  # Colunas desnecessárias

    # remove colunas desnecessárias
    df = df.drop(columns=df_excluir, errors="ignore")  # Remove colunas, ignora se não existir

    return df  # Retorna DataFrame padronizado

#===========================================================================
#   PADRONIZA OS DADOS DO ERP
#===========================================================================
def padronizar_erp(df):  # Função para padronizar o DataFrame do ERP

    # renomeia colunas para um padrão consistente do sistema
    df = df.rename(columns={
        "1o. Agrupamento": "Agrupamento ERP",
        "Ch Criação": "Chave Criação ERP",
        "Chave": "Chave ERP",
        "Pessoa do Título": "Pessoa do Título ERP",
        "Nome do Cliente": "Nome do Cliente ERP",
        "Numero": "Número ERP",
        "NSU": "Código NSU ERP",
        "NSU Concentrador": "NSU Concentrador ERP",
        "Autorização": "Autorização ERP",
        "Emissão": "Data da Transação ERP",
        "Correção": "Correção ERP",
        "Valor": "Valor Bruto ERP",
        "Vr Corrigido": "Valor Líquido ERP",
        "Taxa": "Taxa ERP",
    })  # Renomeia colunas para consistência

    # converte data para datetime padronizado sem horário
    df["Data da Transação ERP"] = pd.to_datetime(
        df["Data da Transação ERP"], dayfirst=True, errors="coerce"
    ).dt.normalize()  # Converte data para datetime sem hora

    # função para converter valores monetários de string para float corretamente
    def converter_valor(col):  # Função interna para converter valores
        return (
            col.astype(str)  # Garante string
            .str.replace("R$", "", regex=False)  # Remove R$
            .str.replace(".", "", regex=False)  # Remove ponto
            .str.replace(",", ".", regex=False)  # Vírgula para ponto
            .astype(float)  # Para float
        )

    # aplica conversão nos valores financeiros
    df["Valor Bruto ERP"] = converter_valor(df["Valor Bruto ERP"])  # Converte valor bruto
    df["Valor Líquido ERP"] = converter_valor(df["Valor Líquido ERP"])  # Converte valor líquido
    df["Taxa ERP"] = converter_valor(df["Taxa ERP"])  # Converte taxa

    # extrai número da parcela e total de parcelas do campo Número ERP
    parcelas = df["Número ERP"].str.extract(r"-(\d+)/(\d+)")  # Extrai parcelas com regex

    # define parcela atual garantindo inteiro e padrão 1 quando vazio
    df["Parcela ERP"] = pd.to_numeric(parcelas[0], errors="coerce").fillna(1).astype(int)  # Parcela atual

    # define total de parcelas garantindo inteiro e padrão 1 quando vazio
    df["Total Parcelas ERP"] = pd.to_numeric(parcelas[1], errors="coerce").fillna(1).astype(int)  # Total parcelas
    
    #remove o sufixo de parcelas do campo Número ERP para manter apenas o número base
    df["Número ERP"] = df["Número ERP"].fillna("").astype(str).str.replace(r"-(\d+)/(\d+)", "", regex=True)

    # garante que Número ERP seja string pois contém texto estruturado
    df["Número ERP"] = df["Número ERP"].astype(str)  # Converte para string

    # função para limpar identificadores removendo problema de ".0" do pandas
    def limpar_id(col):  # Função para limpar IDs
        return (
            pd.to_numeric(col, errors="coerce")  # Para numérico
            .astype("Int64")  # Inteiro com nulos
            .astype(str)  # Para string sem .0
        )

    # padroniza campos chave para evitar erro de conciliação
    df["Código NSU ERP"] = limpar_id(df["Código NSU ERP"])  # Limpa NSU
    df["NSU Concentrador ERP"] = limpar_id(df["NSU Concentrador ERP"])  # Limpa NSU Concentrador
    df["Autorização ERP"] = limpar_id(df["Autorização ERP"])  # Limpa Autorização

    # retorna dataframe tratado e pronto para conciliação
    return df  # Retorna DataFrame padronizado



#===========================================================================
#PROCESSO DE CONCILIAÇÃO ENTRE PAGBANK E ERP
#===========================================================================
def conciliar_pagbank_erp(df_pagbank_padrao, df_erp_padrao, tolerancia_dias=3, tolerancia_valor=0.30):  # Função de conciliação

    df_pagbank_padrao = df_pagbank_padrao.copy()  # Copia para evitar modificações originais
    df_erp_padrao = df_erp_padrao.copy()  # Copia para evitar modificações originais

    # =======
    # LIMPEZA
    # =======
    def limpar_id(valor):  # Função para limpar IDs
        if pd.isna(valor):
            return ""
        return str(valor).replace(".0", "").strip()  # Remove .0 e espaços

    df_erp_padrao["Código NSU ERP"] = df_erp_padrao["Código NSU ERP"].apply(limpar_id)  # Limpa NSU ERP
    df_pagbank_padrao["Código NSU Pagbank"] = df_pagbank_padrao["Código NSU Pagbank"].apply(limpar_id)  # Limpa NSU Pagbank

    df_erp_padrao["Autorização ERP"] = df_erp_padrao["Autorização ERP"].apply(limpar_id)  # Limpa Autorização ERP
    df_pagbank_padrao["Código de Autorização Pagbank"] = df_pagbank_padrao["Código de Autorização Pagbank"].apply(limpar_id)  # Limpa Autorização Pagbank

    df_erp_padrao["Usada"] = False  # Marca se linha ERP foi usada

    # =================
    # COLUNAS RESULTADO
    # =================
    colunas_retorno = [
        "Autorização ERP",
        "Código NSU ERP",
        "Chave ERP",
        "Valor Líquido ERP",
        "Data da Transação ERP",
        "Pessoa do Título ERP"
    ]  # Colunas a adicionar ao resultado

    for col in colunas_retorno:
        df_pagbank_padrao[col] = None  # Inicializa colunas com None

    df_pagbank_padrao["Status"] = "Não conciliado"  # Status inicial
    df_pagbank_padrao["Pontuação"] = 999  # Pontuação inicial alta

    # ============================
    # CLASSIFICAÇÃO DE DIVERGÊNCIA
    # ============================
    def classificar_divergencia(dias, valor, sim_aut, sim_nsu):  # Função para classificar divergências

        problemas = []  # Lista de problemas

        if valor > tolerancia_valor:
            problemas.append("Valor")  # Divergência de valor

        if dias > tolerancia_dias:
            problemas.append("Data")  # Divergência de data

        if sim_aut < 90:
            problemas.append("Autorização")  # Baixa similaridade em autorização

        if sim_nsu < 90:
            problemas.append("NSU")  # Baixa similaridade em NSU

        if not problemas:
            return "OK"  # Sem problemas

        return "Divergência: " + " | ".join(problemas)  # Junta problemas

    # =====
    # SCORE
    # =====
    def calcular_score(linha, row):  # Função para calcular score de similaridade

        dias = abs((linha["Data da Transação ERP"] - row["Data da Transação Pagbank"]).days)  # Diferença em dias
        val = abs(linha["Valor Líquido ERP"] - row["Valor Líquido Pagbank"])  # Diferença em valor

        sim_aut = fuzz.ratio(str(linha["Autorização ERP"]), str(row["Código de Autorização Pagbank"]))  # Similaridade autorização
        sim_nsu = fuzz.ratio(str(linha["Código NSU ERP"]), str(row["Código NSU Pagbank"]))  # Similaridade NSU

        score = (
            dias * 50 +  # Peso para dias
            val * 200 +  # Peso para valor
            (200 - (sim_aut + sim_nsu))  # Peso para dissimilaridade
        )

        return score, dias, val, sim_aut, sim_nsu  # Retorna score e componentes

    # ==============
    # LOOP PRINCIPAL
    # ==============
    for i, row in df_pagbank_padrao.iterrows():  # Itera sobre cada linha do PagBank

        if not row["Código NSU Pagbank"]:  # Se NSU vazio, pula
            continue

        # ============================
        # ETAPA 1 — PERFEITO PAGSEGURO
        # ============================
        match = df_erp_padrao[
            (~df_erp_padrao["Usada"]) &  # Não usada
            (df_erp_padrao["Pessoa do Título ERP"].str.contains("pagseguro", case=False, na=False)) &  # Contém PagSeguro
            (df_erp_padrao["Autorização ERP"] == row["Código de Autorização Pagbank"]) &  # Autorização igual
            (df_erp_padrao["Código NSU ERP"] == row["Código NSU Pagbank"]) &  # NSU igual
            (df_erp_padrao["Valor Líquido ERP"] == row["Valor Líquido Pagbank"]) &  # Valor igual
            (df_erp_padrao["Data da Transação ERP"] == row["Data da Transação Pagbank"])  # Data igual
        ]

        if not match.empty:  # Se encontrou match
            linha = match.iloc[0]  # Primeira linha

            df_erp_padrao.loc[linha.name, "Usada"] = True  # Marca como usada

            df_pagbank_padrao.loc[i, colunas_retorno] = [
                linha["Autorização ERP"],
                linha["Código NSU ERP"],
                linha["Chave ERP"],
                linha["Valor Líquido ERP"],
                linha["Data da Transação ERP"],
                linha["Pessoa do Título ERP"]
            ]  # Preenche colunas

            df_pagbank_padrao.at[i, "Status"] = "Perfeito PagSeguro"  # Status
            df_pagbank_padrao.at[i, "Pontuação"] = 0  # Pontuação 0
            continue  # Próxima iteração

        # ========================
        # ETAPA 2 — PERFEITO GERAL
        # ========================
        match = df_erp_padrao[
            (~df_erp_padrao["Usada"]) &  # Não usada
            (df_erp_padrao["Autorização ERP"] == row["Código de Autorização Pagbank"]) &  # Autorização igual
            (df_erp_padrao["Código NSU ERP"] == row["Código NSU Pagbank"]) &  # NSU igual
            (df_erp_padrao["Valor Líquido ERP"] == row["Valor Líquido Pagbank"]) &  # Valor igual
            (df_erp_padrao["Data da Transação ERP"] == row["Data da Transação Pagbank"])  # Data igual
        ]

        if not match.empty:  # Se encontrou match
            linha = match.iloc[0]  # Primeira linha

            df_erp_padrao.loc[linha.name, "Usada"] = True  # Marca como usada

            df_pagbank_padrao.loc[i, colunas_retorno] = [
                linha["Autorização ERP"],
                linha["Código NSU ERP"],
                linha["Chave ERP"],
                linha["Valor Líquido ERP"],
                linha["Data da Transação ERP"],
                linha["Pessoa do Título ERP"]
            ]  # Preenche colunas

            df_pagbank_padrao.at[i, "Status"] = "Perfeito Geral"  # Status
            df_pagbank_padrao.at[i, "Pontuação"] = 0  # Pontuação 0
            continue  # Próxima iteração

        # ============================
        # ETAPA 3 — TOLERÂNCIA + SCORE
        # ============================
        candidatos = df_erp_padrao[
            (~df_erp_padrao["Usada"]) &  # Não usada
            (abs(df_erp_padrao["Valor Líquido ERP"] - row["Valor Líquido Pagbank"]) <= tolerancia_valor) &  # Valor dentro tolerância
            (abs((df_erp_padrao["Data da Transação ERP"] - row["Data da Transação Pagbank"]).dt.days) <= tolerancia_dias)  # Data dentro tolerância
        ]

        melhores = []  # Lista de melhores candidatos

        for _, linha in candidatos.iterrows():  # Itera candidatos

            score, dias, val, sim_aut, sim_nsu = calcular_score(linha, row)  # Calcula score

            if sim_aut >= 85 or sim_nsu >= 85:  # Se similaridade alta
                melhores.append((linha, score, dias, val, sim_aut, sim_nsu))  # Adiciona à lista

        if melhores:  # Se há melhores
            melhores.sort(key=lambda x: x[1])  # Ordena por score
            linha, score, dias, val, sim_aut, sim_nsu = melhores[0]  # Melhor

            df_erp_padrao.loc[linha.name, "Usada"] = True  # Marca como usada

            df_pagbank_padrao.loc[i, colunas_retorno] = [
                linha["Autorização ERP"],
                linha["Código NSU ERP"],
                linha["Chave ERP"],
                linha["Valor Líquido ERP"],
                linha["Data da Transação ERP"],
                linha["Pessoa do Título ERP"]
            ]  # Preenche colunas

            status = classificar_divergencia(dias, val, sim_aut, sim_nsu)  # Classifica divergência

            df_pagbank_padrao.at[i, "Status"] = status  # Status
            df_pagbank_padrao.at[i, "Pontuação"] = round(score, 0)  # Pontuação arredondada

            continue  # Próxima iteração

    return df_pagbank_padrao, df_erp_padrao  # Retorna DataFrames atualizados


# =========================
# STREAMLIT INTERFACE (PADRÃO CIELO)
# =========================

def main():
    
    st.title("Sistema de Conciliação Bancária")

    # =========================
    # SIDEBAR
    # =========================
    with st.sidebar:
        st.markdown("## 📊 Painel de Controle")

        erp_file = st.file_uploader("ERP (CSV)", type=["csv"])
        pagbank_file = st.file_uploader("PagBank (CSV)", type=["csv"])

        executar = st.button("🚀 Executar Conciliação", use_container_width=True)

    # =========================
    # VALIDAÇÃO
    # =========================
    if not erp_file or not pagbank_file:
        st.warning("⚠️ Envie os arquivos para continuar")
        return

    if executar:

        # =========================
        # CARREGAMENTO
        # =========================
        with st.spinner("Carregando arquivos..."):
            df_erp_padrao = pd.read_csv(erp_file, sep=";", encoding="latin1")
            df_pagbank = pd.read_csv(pagbank_file, sep=";", encoding="utf-8-sig")

        # =========================
        # PADRONIZAÇÃO
        # =========================
        with st.spinner("Padronizando dados..."):
            df_erp_padrao = padronizar_erp(df_erp_padrao)
            df_pagbank_padrao = padronizar_pagbank(df_pagbank)

            
        # =========================
        # CONCILIAÇÃO
        # =========================
        with st.spinner("Conciliando dados..."):
            df_resultado, df_erp_padrao_final = conciliar_pagbank_erp(df_pagbank_padrao, df_erp_padrao)

        # =========================
        # DIVISÃO
        # =========================
        conciliados = df_resultado[df_resultado["Status"] != "Não conciliado"]
        nao_conciliados = df_resultado[df_resultado["Status"] == "Não conciliado"]

        # =========================
        # RESUMO PADRÃO RELATÓRIO
        # =========================
        
        resumo_data = [
            ["RELATÓRIO DE CONCILIAÇÃO", "", ""],

            ["CONCILIADO", "", ""],
            ["- Valor Líquido Total", "", conciliados["Valor Líquido Pagbank"].sum()],
            ["- Valor da Parcela Total", "", conciliados["Valor Bruto Pagbank"].sum()],
            ["- Quantidade de Títulos", "", len(conciliados)],

            ["", "", ""],

            ["NÃO CONCILIADO", "", ""],
            ["- Valor Líquido Total", "", nao_conciliados["Valor Líquido Pagbank"].sum()],
            ["- Valor da Parcela Total", "", nao_conciliados["Valor Bruto Pagbank"].sum()],
            ["- Quantidade de Títulos", "", len(nao_conciliados)],

            ["", "", ""],

            ["CANCELAMENTO DE VENDA", "", ""],
            ["- Valor Líquido Total", "", 0],
            ["- Valor da Parcela Total", "", 0],
            ["- Quantidade de Títulos", "", 0],

            ["", "", ""],

            ["OUTROS", "", ""],
            ["- Valor total de aluguel de maquineta", "", 0],
            ["- Valor Total no Banco", "", df_resultado["Valor Líquido Pagbank"].sum()],
        ]

        resumo = pd.DataFrame(resumo_data, columns=["Categoria", "Descrição", "Valor"])



        # =========================
        # KPIs
        # =========================
        st.markdown("## 📊 Resultado")

        col1, col2 = st.columns(2)
        col1.metric("Conciliados", len(conciliados))
        col2.metric("Não conciliados", len(nao_conciliados))

        # =========================
        # TABS
        # =========================
        tab1, tab2, tab3 = st.tabs(["Conciliados", "Não conciliados", "Resumo"])

        with tab1:
            st.dataframe(conciliados, use_container_width=True)

        with tab2:
            st.dataframe(nao_conciliados, use_container_width=True)

        with tab3:
            st.dataframe(resumo, use_container_width=True)

    # =========================
    # DOWNLOAD
    # =========================
    if executar:

        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            output.seek(0)

            conciliados.to_excel(writer, sheet_name="Conciliados", index=False)
            nao_conciliados.to_excel(writer, sheet_name="Nao_Conciliados", index=False)
            df_resultado.to_excel(writer, sheet_name="Completo", index=False)
            resumo.to_excel(writer, sheet_name="Resumo", index=False)

            workbook = writer.book
            ws_conciliados = writer.sheets["Conciliados"]
            ws_resumo = writer.sheets["Resumo"]

            # =========================
            # EXTRAIR CHAVES ERP
            # =========================
            header = [cell.value for cell in ws_conciliados[1]]

            if "Chave ERP" in header:
                idx_chave = header.index("Chave ERP")

                chaves = [
                    str(row[idx_chave]).strip()
                    for row in ws_conciliados.iter_rows(min_row=2, values_only=True)
                    if row[idx_chave] not in (None, "", "nan")
                ]

                blocos = []
                bloco_atual = []
                tamanho_atual = 0

                for chave in chaves:
                    if tamanho_atual + len(chave) + 2 > 30000:
                        blocos.append(bloco_atual)
                        bloco_atual = []
                        tamanho_atual = 0

                    bloco_atual.append(chave)
                    tamanho_atual += len(chave) + 2

                if bloco_atual:
                    blocos.append(bloco_atual)

                start_row = ws_resumo.max_row + 2

                for i, bloco in enumerate(blocos, start=1):
                    texto = ", ".join(bloco)

                    ws_resumo.cell(row=start_row + i - 1, column=1, value=f"Grupo {i}")
                    ws_resumo.cell(row=start_row + i - 1, column=2, value=texto)

            # =========================
            # FORMATAÇÃO
            # =========================
            bold = Font(bold=True)

            for row in ws_resumo.iter_rows():
                categoria = row[0].value

                if categoria in [
                    "RELATÓRIO DE CONCILIAÇÃO",
                    "CONCILIADO",
                    "NÃO CONCILIADO",
                    "CANCELAMENTO DE VENDA",
                    "OUTROS"
                ]:
                    row[0].font = bold

        valor_cell = row[2]

        try:
            valor = float(valor_cell.value)
            valor_cell.value = valor  # garante número real
            valor_cell.number_format = 'R$ #,##0.00'
        except (TypeError, ValueError):
            pass  # ignora valores inválidos

        # =========================
        # BOTÃO DOWNLOAD
        # =========================
        st.markdown("---")
        st.download_button(
            "📥 Baixar Excel",
            data=output.getvalue(),
            file_name="conciliacao_pagbank.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


if __name__ == "__main__":
    main()